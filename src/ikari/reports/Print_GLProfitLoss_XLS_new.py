import calendar
import datetime
import re
from openpyxl import load_workbook
from decimal import Decimal
import xlsxwriter
from django.core.files.base import ContentFile
from django.db.models import Q
from accounting.models import FiscalCalendar
from accounts.models import Account, AccountHistory, ReportGroup
from companies.models import Company, CostCenters
from reports.print_GLBalance_Sheet_XLS_new import get_result
from reports.print_GLProfit_Loss_new import getProvisionalPostedAmount
from utilities.common import round_number
from utilities.constants import REPORT_TYPE
from django.contrib.humanize.templatetags.humanize import intcomma


class Print_GLProfitLoss_XLS_new:
    def __init__(self, buffer):
        self.buffer = buffer
        self.bottom_style = ''
        self.top_style = ''
    
    def set_cell_style(self, cell):
        if cell.has_style:
            if cell.border.bottom.border_style == 'double' or \
                cell.border.bottom.border_style == 'thin':
                self.bottom_style = cell.border.bottom.border_style
            else:
                self.bottom_style = ''
            if cell.border.top.border_style == 'double' or \
                cell.border.top.border_style == 'thin':
                self.top_style = cell.border.top.border_style
            else:
                self.top_style = ''
        else:
            self.top_style = ''
            self.bottom_style = ''
        

    def WriteToExcel(self, company_id, issue_date, report_type, filter_type, from_val, to_val):
        company = Company.objects.get(pk=company_id)
        output = self.buffer
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet("Profit_Loss")
        printing_row = 10
        printing_col = 0
        merge_format = workbook.add_format({
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
            # 'bg_color': 'white'
        })
        merge_str = workbook.add_format({
            'valign': 'vcenter',
        })
        right_line = workbook.add_format({
            'align': 'right'
        })
        center = workbook.add_format({
            'align': 'center'
        })
        center_line = workbook.add_format({
            'align': 'center',
            'bottom': 1
        })
        # border_top_bot = workbook.add_format({
        #     # 'bold': True,
        #     'align': 'right',
        #     'bottom': 1,
        #     'top': 1
        # })

        # border_top = workbook.add_format({
        #     'align': 'right',
        #     'top': 1
        # })

        dec_format = workbook.add_format({
            'num_format': '#,##0.00_);(#,##0.00)',
            'align': 'right',
        })
        num_format = workbook.add_format({
            'num_format': '#,##0_);(#,##0)',
            'align': 'right',
        })

        array_data = str(issue_date).split('-')
        month = int(array_data[1])
        year = int(array_data[0])
        issue_from = datetime.date(int(array_data[0]), int(array_data[1]), 1)
        issue_to = datetime.date(year, month,
                                 calendar.monthrange(year, month)[1])
        company = Company.objects.get(pk=company_id)
        segment_ids = []
        segment_name = ''
        if filter_type == 'undefined' or filter_type == '' or filter_type == 'None':
            filter_type = 'Account'
        if filter_type == 'Segment':
            try:
                from_val = int(from_val)
                to_val = int(to_val)
                segments = CostCenters.objects.filter(id__range=(from_val, to_val))
                for seg in segments:
                    segment_name += ' - ' + seg.name
                segment_ids = segments.values_list('id', flat=True)
            except:
                pass
        fsc_calendar = FiscalCalendar.objects.filter(company_id=company_id, is_hidden=0, period=month, fiscal_year=year).last()
        if fsc_calendar:
            end_date = fsc_calendar.end_date.strftime('%d/%m/%Y')
        else:
            end_date = issue_to.strftime('%d/%m/%Y')

        worksheet.merge_range('A3:C3', str(company.name) + str(segment_name), merge_format)
        worksheet.merge_range('A4:C4', 'TRADING AND PROFIT AND LOSS ACCOUNT', merge_format)
        worksheet.merge_range('A5:C5', 'FOR THE MONTH OF ' + end_date, merge_format)
        worksheet.write(6, 2, 'CURRENT', center)
        worksheet.write(6, 3, 'YEAR TO', center)
        worksheet.write(7, 2, 'MONTH', center)
        worksheet.write(7, 3, 'DATE', center)
        worksheet.write(8, 2, company.currency.code, center_line)
        worksheet.write(8, 3, company.currency.code, center_line)
        worksheet.set_column(0, 0, 12)
        worksheet.set_column(1, 1, 25)
        worksheet.set_column(2, 3, 16)

        decimal_place = "%.2f"
        is_decimal = True
        if not company.currency.is_decimal:
            decimal_place = "%.0f"
            is_decimal = False

        accountHistory = AccountHistory.objects.select_related('account').filter(company_id=company.id, is_hidden=0,
                                                                                 period_month__exact=month,
                                                                                 period_year__exact=year)\
            .exclude(source_currency_id__isnull=True)

        if company.pl_template:
            try:
                bytefile = company.pl_template.read()
                wb = load_workbook(ContentFile(bytefile))
                sheet_0 = wb.get_sheet_names()[0]
                ws = wb[sheet_0]

                hidden_rows = []
                for rowNum, rowDimension in ws.row_dimensions.items():
                    if rowDimension.hidden == True:
                        hidden_rows.append(rowNum)
                hidden_cols = []
                for colNum, colDimension in ws.column_dimensions.items():
                    if colDimension.hidden == True:
                        hidden_cols.append(colNum)

                acc_pattern_1 = '[=\"][A-Z]*[0-9]+[:][A-Z]*[0-9]+[Z]*[\"]'
                acc_pattern_2 = '[=\"][A-Z]*[0-9]+[%]*[\"]'
                sum_pattern_1 = re.compile(r'SUM')
                sum_pattern_2 = '[A-Z][0-9]+[+]*[-]*[A-Z]*[0-9]*'
                sum_pattern_3 = '[\\\\][-]*'
                seg_codes = list(CostCenters.objects.filter(is_active=1, is_hidden=0, company_id=company_id).values_list('code', flat=True))
                account_item_list = Account.objects.filter(company_id=company.id, is_hidden=0).exclude(deactivate_period__lte=issue_from)
                account_item_list = account_item_list.exclude(Q(is_active=False) & Q(deactivate_date=None))
                change_array = [0] * (ws.max_row + 1)
                balance_array = [0] * (ws.max_row + 1)
                title_pattern = '[a-zA-Z]+[,]*[a-zA-Z]+[:]*[^0-9]+'
                start = False
                try:
                    for row in range(1, ws.max_row + 5):
                        empty_row = True
                        if row not in hidden_rows:
                            for col in range(1, ws.max_column + 5):
                                if col not in hidden_cols:
                                    if ws.cell(row, col).value is not None:
                                        if start:
                                            if re.search(acc_pattern_1, str(ws.cell(row, col).value)):
                                                empty_row = False
                                                total_change = 0
                                                total_balance = 0
                                                acc_range = str(ws.cell(row, col).value)
                                                acc_range = acc_range.replace('Z', '9').replace('=', '').replace('"', '').strip().split(':')
                                                f_code = int(re.sub('[A-Z]+', '', acc_range[0]))
                                                l_code = int(re.sub('[A-Z]+', '', acc_range[1]))
                                                code_list = []
                                                while f_code <= l_code:
                                                    code_list.append(str(f_code))
                                                    f_code += 1
                                                account_list = account_item_list.filter(account_segment__in=code_list).order_by('account_segment', 'code')
                                                if filter_type == 'Segment':
                                                    account_list = account_list.filter(segment_code_id__in=segment_ids)
                                                code_list = account_list.order_by('account_segment').values_list('account_segment', flat=True).distinct()
                                                col += 1
                                                while not re.search(sum_pattern_3, str(ws.cell(row, col).value)) and col < ws.max_column:
                                                    col += 1
                                                sign = Decimal(1.0)
                                                if re.search('\-', str(ws.cell(row, col).value)):
                                                    sign = Decimal(-1.0)
                                                for cod in code_list:
                                                    acc_total_change = 0
                                                    acc_total_balance = 0
                                                    segments = account_list.filter(account_segment=cod)
                                                    for f_acc in segments:
                                                        item_account_q = accountHistory.filter(account_id=f_acc.id)
                                                        item_account = item_account_q.filter(
                                                            Q(source_currency_id=company.currency_id) |
                                                            Q(source_currency_id__isnull=True))
                                                        sum_change = sum_balance = 0
                                                        if item_account:
                                                            prov_posted_amount_per_account = 0
                                                            for iAccount in item_account:
                                                                if report_type == dict(REPORT_TYPE)['Provisional']:
                                                                    prov_posted_amount_per_account = getProvisionalPostedAmount(company.id, year, month,
                                                                                                                                f_acc.id,
                                                                                                                                iAccount.source_currency_id)
                                                                sum_change += (iAccount.functional_net_change * sign) + Decimal(prov_posted_amount_per_account)
                                                                sum_balance += (iAccount.functional_end_balance * sign) + \
                                                                    Decimal(prov_posted_amount_per_account)
                                                                acc_total_change += (iAccount.functional_net_change * sign) + \
                                                                    Decimal(prov_posted_amount_per_account)
                                                                acc_total_balance += (iAccount.functional_end_balance * sign) + \
                                                                    Decimal(prov_posted_amount_per_account)
                                                                total_change += (iAccount.functional_net_change * sign) + \
                                                                    Decimal(prov_posted_amount_per_account)
                                                                total_balance += (iAccount.functional_end_balance * sign) + \
                                                                    Decimal(prov_posted_amount_per_account)
                                                            if filter_type == 'Segment':
                                                                worksheet.write(printing_row, printing_col, f_acc.code)
                                                                worksheet.write(printing_row, printing_col + 1, f_acc.name)
                                                                worksheet.write(printing_row, printing_col + 2,
                                                                                wrap_text_xls(sum_change, decimal_place), dec_format if is_decimal else num_format)
                                                                worksheet.write(printing_row, printing_col + 3,
                                                                                wrap_text_xls(sum_balance, decimal_place), dec_format if is_decimal else num_format)
                                                                printing_row += 1
                                                    if filter_type == 'Account':
                                                        f_acc = account_item_list.filter(account_segment=str(cod)).first()
                                                        if f_acc:
                                                            code = f_acc.code.split('-')[0]
                                                            if company.use_segment:
                                                                names = f_acc.name.split('-')
                                                                for cd in seg_codes:
                                                                    if names[-1] in cd:
                                                                        names.pop()
                                                                name = '-'.join(names)
                                                            else:
                                                                name = f_acc.name
                                                            worksheet.write(printing_row, printing_col, code)
                                                            worksheet.write(printing_row, printing_col + 1, name)
                                                            worksheet.write(printing_row, printing_col + 2,
                                                                            wrap_text_xls(acc_total_change, decimal_place), dec_format if is_decimal else num_format)
                                                            worksheet.write(printing_row, printing_col + 3,
                                                                            wrap_text_xls(acc_total_balance, decimal_place), dec_format if is_decimal else num_format)
                                                            printing_row += 1
                                                change_array[row] = float(total_change)
                                                balance_array[row] = float(total_balance)
                                                break
                                            elif re.search(acc_pattern_2, str(ws.cell(row, col).value)):
                                                empty_row = False
                                                total_change = 0
                                                total_balance = 0
                                                f_code = str(ws.cell(row, col).value).replace('%', '').replace('=', '').replace('"', '').strip()
                                                f_code = re.sub('[A-Z]+', '', f_code)
                                                acc_list = account_item_list.filter(account_segment=str(f_code))
                                                if filter_type == 'Segment':
                                                    acc_list = acc_list.filter(segment_code_id__in=segment_ids)
                                                acc_total_change = 0
                                                acc_total_balance = 0
                                                col += 1
                                                while ws.cell(row, col).value is None or str(ws.cell(row, col).value) == 'D(ACCT)':
                                                    col += 1
                                                name_str = str(ws.cell(row, col).value)

                                                col += 1
                                                while not re.search(sum_pattern_3, str(ws.cell(row, col).value)) and col < ws.max_column:
                                                    col += 1
                                                sign = Decimal(1.0)
                                                if re.search('\-', str(ws.cell(row, col).value)):
                                                    sign = Decimal(-1.0)
                                                for f_acc in acc_list:
                                                    item_account_q = accountHistory.filter(account_id=f_acc.id)
                                                    item_account = item_account_q.filter(
                                                        Q(source_currency_id=company.currency_id) |
                                                        Q(source_currency_id__isnull=True))
                                                    sum_change = sum_balance = 0
                                                    if item_account:
                                                        prov_posted_amount_per_account = 0
                                                        for iAccount in item_account:
                                                            if report_type == dict(REPORT_TYPE)['Provisional']:
                                                                prov_posted_amount_per_account = getProvisionalPostedAmount(company.id, year, month,
                                                                                                                            f_acc.id,
                                                                                                                            iAccount.source_currency_id)
                                                            sum_change += (iAccount.functional_net_change * sign) + Decimal(prov_posted_amount_per_account)
                                                            sum_balance += (iAccount.functional_end_balance * sign) + Decimal(prov_posted_amount_per_account)
                                                            acc_total_change += (iAccount.functional_net_change * sign) + \
                                                                Decimal(prov_posted_amount_per_account)
                                                            acc_total_balance += (iAccount.functional_end_balance * sign) + \
                                                                Decimal(prov_posted_amount_per_account)
                                                            total_change += (iAccount.functional_net_change * sign) + Decimal(prov_posted_amount_per_account)
                                                            total_balance += (iAccount.functional_end_balance * sign) + Decimal(prov_posted_amount_per_account)
                                                        if filter_type == 'Segment':
                                                            if re.search('ACCTDESC"\)$', name_str):
                                                                code = str(f_acc.code)
                                                                name = str(f_acc.name)
                                                            elif re.search(str(f_acc.account_segment), name_str):
                                                                code = str(f_acc.code)
                                                                name = re.sub(str(f_acc.account_segment), '', name_str)
                                                                name = name.replace('&', '').replace('=', '').replace('"', '').strip()
                                                            else:
                                                                code = str(f_acc.code)
                                                                name = str(f_acc.name)

                                                            worksheet.write(printing_row, printing_col, code)
                                                            worksheet.write(printing_row, printing_col + 1, name)
                                                            worksheet.write(printing_row, printing_col + 2,
                                                                            wrap_text_xls(sum_change, decimal_place), dec_format if is_decimal else num_format)
                                                            worksheet.write(printing_row, printing_col + 3,
                                                                            wrap_text_xls(sum_balance, decimal_place), dec_format if is_decimal else num_format)
                                                            printing_row += 1
                                                change_array[row] = float(total_change)
                                                balance_array[row] = float(total_balance)
                                                if filter_type == 'Account':
                                                    f_acc = account_item_list.filter(account_segment=str(f_code)).first()
                                                    if f_acc:
                                                        code = f_acc.code.split('-')[0]
                                                        if company.use_segment:
                                                            names = f_acc.name.split('-')
                                                            for cd in seg_codes:
                                                                if names[-1] in cd:
                                                                    names.pop()
                                                            name = '-'.join(names)
                                                        else:
                                                            name = f_acc.name
                                                        worksheet.write(printing_row, printing_col, code)
                                                        worksheet.write(printing_row, printing_col + 1, name)
                                                        worksheet.write(printing_row, printing_col + 2,
                                                                        wrap_text_xls(acc_total_change, decimal_place), dec_format if is_decimal else num_format)
                                                        worksheet.write(printing_row, printing_col + 3,
                                                                        wrap_text_xls(acc_total_balance, decimal_place), dec_format if is_decimal else num_format)
                                                        printing_row += 1
                                                break
                                            elif re.search(title_pattern, str(ws.cell(row, col).value).replace(' ', '').lower()) or \
                                                    re.search(sum_pattern_1, str(ws.cell(row, col).value)) or re.search(sum_pattern_2, str(ws.cell(row, col).value)):
                                                empty_row = False
                                                if re.search(sum_pattern_1, str(ws.cell(row, col).value)) or re.search(sum_pattern_2, str(ws.cell(row, col).value)):
                                                    change = get_result(str(ws.cell(row, col).value), change_array)
                                                    col += 1
                                                    while ws.cell(row, col).value is None and col < ws.max_column:
                                                        col += 1
                                                    balance = get_result(str(ws.cell(row, col).value), balance_array)
                                                    self.set_cell_style((ws.cell(row, col)))
                                                    if is_decimal:
                                                        local_border = workbook.add_format({
                                                            'num_format': '#,##0.00_);(#,##0.00)',
                                                            'align': 'right'
                                                        })
                                                    else:
                                                        local_border = workbook.add_format({
                                                            'num_format': '#,##0_);(#,##0)',
                                                            'align': 'right'
                                                        })
                                                    if self.top_style == 'thin':
                                                        local_border.set_top(1)
                                                    elif self.top_style == 'double':
                                                        local_border.set_top(6)
                                                    if self.bottom_style == 'thin':
                                                        local_border.set_bottom(1)
                                                    elif self.bottom_style == 'double':
                                                        local_border.set_bottom(6)
                                                    
                                                    worksheet.write(printing_row, printing_col + 2, wrap_text_xls(change, decimal_place), local_border)
                                                    worksheet.write(printing_row, printing_col + 3, wrap_text_xls(balance, decimal_place), local_border)
                                                    printing_row += 1

                                                    change_array[row] = float(change)
                                                    balance_array[row] = float(balance)

                                                else:
                                                    total_str = str(ws.cell(row, col).value).replace('=', '').replace('"', '')
                                                    col += 1
                                                    while ws.cell(row, col).value is None and col < ws.max_column:
                                                        col += 1
                                                    if col < ws.max_column:
                                                        if re.search(sum_pattern_1, str(ws.cell(row, col).value)) or re.search(sum_pattern_2, str(ws.cell(row, col).value)):
                                                            change = get_result(str(ws.cell(row, col).value), change_array)
                                                            col += 1
                                                            while ws.cell(row, col).value is None and col < ws.max_column:
                                                                col += 1
                                                            balance = get_result(str(ws.cell(row, col).value), balance_array)
                                                            self.set_cell_style((ws.cell(row, col)))
                                                            if is_decimal:
                                                                local_border = workbook.add_format({
                                                                    'num_format': '#,##0.00_);(#,##0.00)',
                                                                    'align': 'right'
                                                                })
                                                            else:
                                                                local_border = workbook.add_format({
                                                                    'num_format': '#,##0_);(#,##0)',
                                                                    'align': 'right'
                                                                })
                                                            if self.top_style == 'thin':
                                                                local_border.set_top(1)
                                                            elif self.top_style == 'double':
                                                                local_border.set_top(6)
                                                            if self.bottom_style == 'thin':
                                                                local_border.set_bottom(1)
                                                            elif self.bottom_style == 'double':
                                                                local_border.set_bottom(6)

                                                            worksheet.merge_range('A' + str(printing_row + 1) + ':' + 'B' + str(printing_row + 1), total_str, merge_str)
                                                            worksheet.write(printing_row, printing_col + 2,
                                                                            wrap_text_xls(change, decimal_place), local_border)
                                                            worksheet.write(printing_row, printing_col + 3,
                                                                            wrap_text_xls(balance, decimal_place), local_border)
                                                            printing_row += 1

                                                            change_array[row] = float(change)
                                                            balance_array[row] = float(balance)

                                                    else:
                                                        worksheet.merge_range('A' + str(printing_row + 1) + ':' + 'B' + str(printing_row + 1), total_str, merge_str)
                                                        printing_row += 1
                                                break
                                        else:
                                            empty_row = False
                                        if re.search(str(company.currency.code).lower(), str(ws.cell(row, col).value).lower()):
                                            start = True
                                            empty_row = False
                                            printing_row += 1
                                            break
                            if empty_row:
                                printing_row += 1

                except Exception as e:
                    print(e)
                    worksheet.write(printing_row, printing_col, 'ERROR! ON FILE PROCESSING')
                    printing_row += 1
                    worksheet.write(printing_row, printing_col, 'PLEASE MAKE SURE IT IS CORRECT TEMPLATE.')
                    printing_row += 1

            except Exception as e:
                print(e)
                worksheet.write(printing_row, printing_col, 'ERROR! CAN NOT READ THE TEMPLATE.')
                printing_row += 1

        else:  # No template
            worksheet.write(printing_row, printing_col, 'ERROR! THERE IS NO TEMPLATE TO PROCESS.')
            printing_row += 1

        workbook.close()
        xlsx_data = output.getvalue()
        return xlsx_data


def wrap_text_xls(value, decimal_place="%.2f"):
    str_value = intcomma(decimal_place % round_number(value)).replace("-", "")
    if value < 0:
        str_value = '(' + str_value + ')'

    return str_value


def createCustomAccGroupList(company_id):
    rpt_acc_grp_list = []
    rpt_acct_grp = ReportGroup.objects.filter(company_id=company_id, is_hidden=False, report_template_type='0')
    for rpt_acc_grp in rpt_acct_grp:
        rpt_acc_grp_obj = {'acc1_id': rpt_acc_grp.account_from.id,
                           'acc2_id': rpt_acc_grp.account_to.id if rpt_acc_grp.account_to else rpt_acc_grp.account_from.id,
                           'acc_code': rpt_acc_grp.account_code_text,
                           'acc_name': rpt_acc_grp.name}
        rpt_acc_grp_list.append(rpt_acc_grp_obj)
    return rpt_acc_grp_list


def getCustomAccGroup(rpt_acc_grp_list, mAccount):
    is_4550_exist = False
    custom_acc = {"custom_acc_code": None, "custom_acc_name": None}
    for custom_acc in rpt_acc_grp_list:
        if mAccount.id >= custom_acc['acc1_id'] and mAccount.id <= custom_acc['acc2_id']:
            custom_acc['custom_acc_code'] = custom_acc['acc_code']
            custom_acc['custom_acc_name'] = str(custom_acc['acc_name'])
            if mAccount.code == '4550':
                is_4550_exist = True
            break
        else:
            custom_acc['custom_acc_code'] = mAccount.code
            custom_acc['custom_acc_name'] = str(mAccount.name)
    return custom_acc, is_4550_exist
