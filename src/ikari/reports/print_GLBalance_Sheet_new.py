import calendar
import datetime
import os
import re
from openpyxl import load_workbook
from decimal import Decimal
from functools import partial
from django.db.models import Q
from django.conf import settings as s
from django.core.files.base import ContentFile
from reportlab.lib import colors
from reportlab.lib.enums import TA_RIGHT, TA_LEFT
from reportlab.lib.pagesizes import A4, letter, portrait
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from accounting.models import FiscalCalendar
from accounts.models import Account, AccountHistory, ReportGroup
from companies.models import Company, CostCenters
from reports.numbered_page import NumberedPage
from reports.print_GLProfit_Loss_new import wrap_text, getProvisionalPostedAmount
from reports.print_GLBalance_Sheet_XLS_new import get_result
from utilities.constants import BALANCE_TYPE_DICT, REPORT_TYPE, REPORT_TEMPLATE_TYPES_DICT

acc_pattern_1 = '[=\"][A-Z]*[0-9]+[:][A-Z]*[0-9]+[Z]*[\"]'
acc_pattern_2 = '[=\"][A-Z]*[0-9]+[%]*[\"]'
acc_pattern_3 = '[=\"]*[A-Z]*[0-9]+[%]*[,][A-Z]*[0-9]+[%]*[\"]*'
acc_pattern_4 = '[=\"]*[A-Z]*[0-9]+[:][A-Z]*[0-9]+[Z]*[,][A-Z]*[0-9]+[%]*[\"]*'
accumulate_pattern = '[0-9]+[-][0-9]+'
sum_pattern_1 = re.compile(r'SUM')
sum_pattern_2 = '[G-H][0-9]+[+]*[-]*[G-H]*[0-9]*'
sum_pattern_3 = '[\\\\][-]*'
sum_pattern_4 = '[0-9]+[:][0-9]+'
code_pattern_1 = '[_xll\.]*FRACCT\(\"ACCTID\"\)'
code_pattern_2 = '[_xll\.]*FRACCT\(\"ACCTFMTTD\"\)'
name_pattern = '[_xll\.]*FRACCT\(\"ACCTDESC\"\)'

styles = getSampleStyleSheet()
styles.add(ParagraphStyle(name='RightAlignBold', fontName=s.REPORT_FONT_BOLD, alignment=TA_RIGHT, fontSize=s.REPORT_FONT_SIZE))
styles.add(ParagraphStyle(name='LeftAlignBold', fontName=s.REPORT_FONT_BOLD, alignment=TA_LEFT, fontSize=s.REPORT_FONT_SIZE))
styles.add(ParagraphStyle(name='RightAlign', fontName=s.REPORT_FONT_BOLD, textColor=colors.red, alignment=TA_RIGHT, fontSize=s.REPORT_FONT_SIZE))
styles.add(ParagraphStyle(name='RightAlignBlack', fontName=s.REPORT_FONT, alignment=TA_RIGHT, fontSize=s.REPORT_FONT_SIZE))
styles.add(ParagraphStyle(name='RightAlignRed', fontName=s.REPORT_FONT, textColor=colors.black, alignment=TA_RIGHT, fontSize=s.REPORT_FONT_SIZE))


class Print_GLBalanceSheet_new:
    def __init__(self, buffer, pagesize, report_type):
        self.buffer = buffer
        if pagesize == 'A4':
            self.pagesize = A4
        elif pagesize == 'Letter':
            self.pagesize = letter
        self.width, self.height = self.pagesize
        self.report_type = report_type
        self.bottom_style = ''
        self.top_style = ''

    @staticmethod
    def _header_footer(canvas, doc, company_id, issue_from, issue_to):
        pdfmetrics.registerFont(TTFont('Arial', os.path.join(s.BASE_DIR, "static/fonts/arial.ttf")))
        pdfmetrics.registerFont(TTFont('Arial-Bold', os.path.join(s.BASE_DIR, "static/fonts/arial-bold.ttf")))

        # Save the state of our canvas so we can draw on it
        canvas.saveState()
        # Draw header of PDF
        # 1st row
        header_data = []
        row1_info1 = datetime.datetime.now().strftime('%d/%m/%Y %r')
        row1_info2 = ""
        header_data.append([row1_info1, row1_info2])

        header_table = Table(header_data, colWidths=[280, 200, 200])
        header_table.setStyle(TableStyle(
            [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
             ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
             ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
             ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
             ('TOPPADDING', (0, 0), (-1, -1), -0.2),
             ('BOTTOMPADDING', (0, 0), (-1, -1), -0.2),
             ]))

        w, h = header_table.wrap(doc.width, doc.topMargin)
        header_table.drawOn(canvas, doc.leftMargin - 7, doc.height + doc.topMargin - h)
        # Release the canvas
        canvas.restoreState()

    def print_report(self, company_id, issue_from, filter_type, from_val, to_val):
        pdfmetrics.registerFont(TTFont('Arial', os.path.join(s.BASE_DIR, "static/fonts/arial.ttf")))
        pdfmetrics.registerFont(TTFont('Arial-Bold', os.path.join(s.BASE_DIR, "static/fonts/arial-bold.ttf")))

        buffer = self.buffer
        self.pagesize = portrait(A4)
        doc = SimpleDocTemplate(buffer, rightMargin=s.REPORT_RIGHT_MARGIN, leftMargin=s.REPORT_LEFT_MARGIN, topMargin=s.REPORT_TOP_MARGIN - 60,
                                bottomMargin=s.REPORT_BOTTOM_MARGIN, pagesize=portrait(A4))
        # Draw Content of PDF
        # design interface header
        # condition input
        array_data = str(issue_from).split('-')
        month = int(array_data[1])
        year = int(array_data[0])
        issue_from = datetime.date(int(array_data[0]), int(array_data[1]), 1)
        issue_to = datetime.date(int(array_data[0]), int(array_data[1]),
                                 calendar.monthrange(int(array_data[0]), int(array_data[1]))[1])
        self.company_id = company_id
        company = Company.objects.get(pk=company_id)
        fsc_calendar = FiscalCalendar.objects.filter(company_id=company_id, is_hidden=0, period=month, fiscal_year=year).last()
        if fsc_calendar:
            end_date = fsc_calendar.end_date.strftime('%d/%m/%Y')
        else:
            end_date = issue_to.strftime('%d/%m/%Y')

        elements = []

        accountHistory = AccountHistory.objects.select_related('account').filter(company_id=company.id, is_hidden=0,
                                                                                 period_month__exact=month,
                                                                                 period_year__exact=year)\
            .exclude(source_currency_id__isnull=True)

        acc_colWidths = [210, 70, 230, 50, 100, 200]
        seg_colWidths = [180, 100, 230, 50, 100, 200]
        rowHeights = 14
        colWidths = acc_colWidths

        segment_ids = []
        segment_name = ''
        if filter_type == 'undefined' or filter_type == '' or filter_type == 'None':
            filter_type = 'Account'
        if filter_type == 'Segment':
            colWidths = seg_colWidths
            try:
                from_val = int(from_val)
                to_val = int(to_val)
                segments = CostCenters.objects.filter(id__range=(from_val, to_val))
                for seg in segments:
                    segment_name += ' - ' + seg.name
                segment_ids = segments.values_list('id', flat=True)
            except:
                pass

        decimal_place = "%.2f"
        if not company.currency.is_decimal:
            decimal_place = "%.0f"

        if company.bs_template:
            try:
                bytefile = company.bs_template.read()
                wb = load_workbook(ContentFile(bytefile))
                sheet_0 = wb.get_sheet_names()[0]
                ws = wb[sheet_0]

                hidden_rows = []
                for rowNum, rowDimension in ws.row_dimensions.items():
                    if rowDimension.hidden == True:
                        hidden_rows.append(rowNum)
                hidden_cols = []
                for colNum, colDimension in ws.column_dimensions.items():
                    if colDimension.hidden == True:
                        hidden_cols.append(colNum)

                start = False
                negative_start = False
                total_account = 0
                seg_codes = list(CostCenters.objects.filter(is_active=1, is_hidden=0, company_id=company_id).values_list('code', flat=True))
                account_item_list = Account.objects.filter(company_id=company.id, is_hidden=0).exclude(deactivate_period__lte=issue_from)
                account_item_list = account_item_list.exclude(Q(is_active=False) & Q(deactivate_date=None))
                value_array = [0] * (ws.max_row + 5)
                title_pattern = '[a-zA-Z]+[,]*[a-zA-Z]+[:]*[^0-9]+'
                try:
                    for row in range(1, ws.max_row + 1):
                        empty_row = True
                        if row not in hidden_rows:
                            for col in range(1, ws.max_column + 5):
                                if col not in hidden_cols:
                                    if ws.cell(row, col).value is not None:
                                        if re.search('BALP'.lower(), str(ws.cell(row, col).value).lower()):
                                            break
                                        if re.search('CONAME'.lower(), str(ws.cell(row, col).value).lower()):
                                            table_data = []
                                            table_data.append(['', str(company.name) + str(segment_name), ''])
                                            item_table = Table(table_data, colWidths=[100, 310, 100])
                                            item_table.setStyle(TableStyle(
                                                [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                 ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                 ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                 ('FONTSIZE', (0, 0), (-1, -1), 12),
                                                 ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                                                 ]))
                                            elements.append(item_table)
                                            break
                                        elif re.search('BALANCE SHEET'.lower(), str(ws.cell(row, col).value).lower()):
                                            table_data = []
                                            table_data.append(['', 'BALANCE SHEET', ''])
                                            item_table = Table(table_data, colWidths=[100, 310, 100])
                                            item_table.setStyle(TableStyle(
                                                [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                 ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                 ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                 ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                                                 ]))
                                            elements.append(item_table)
                                            break
                                        elif re.search('AS AT'.lower(), str(ws.cell(row, col).value).lower()):
                                            table_data = []
                                            table_data.append(['', "AS AT " + end_date, ''])
                                            item_table = Table(table_data, colWidths=[100, 310, 100])
                                            item_table.setStyle(TableStyle(
                                                [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                 ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                 ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                 ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                                                 ]))
                                            elements.append(item_table)
                                            break
                                        elif re.search(str(company.currency.code).lower(), str(ws.cell(row, col).value).lower()):
                                            start = True
                                            table_data = []
                                            table_data.append(
                                                ['', '', '', '', '', ''])
                                            table_data.append(
                                                ['', '', '', '', str(company.currency.code), ''])
                                            item_table = Table(table_data, colWidths=colWidths)
                                            item_table.setStyle(TableStyle(
                                                [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                 ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                 ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                 ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
                                                 ('LEFTPADDING', (0, 0), (-1, -1), 0),
                                                 ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                                                 ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                                 ('LINEBELOW', (4, 1), (4, 1), 0.25, colors.black),
                                                 ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                                                 ]))
                                            elements.append(item_table)
                                            break
                                        if start:
                                            if re.search(title_pattern, str(ws.cell(row, col).value).replace(' ', '').lower()) or \
                                                    re.search(sum_pattern_1, str(ws.cell(row, col).value)) or re.search(sum_pattern_2, str(ws.cell(row, col).value)):
                                                empty_row = False
                                                if re.search(sum_pattern_1, str(ws.cell(row, col).value)) or re.search(sum_pattern_2, str(ws.cell(row, col).value)):
                                                    result = get_result(str(ws.cell(row, col).value), value_array)
                                                    self.set_cell_style((ws.cell(row, col)))
                                                    value_array[row] = float(result)
                                                    elements = self.total_style(elements, '', result, colWidths, decimal_place)
                                                elif re.search(title_pattern, str(ws.cell(row, col).value).replace(' ', '').lower()):
                                                    if re.search('FIXEDASSETS,ATCOST'.lower(), str(ws.cell(row, col).value).replace(' ', '').lower()):
                                                        negative_start = False
                                                    if re.search('INVESTMENT'.lower(), str(ws.cell(row, col).value).replace(' ', '').lower()) or \
                                                            re.search('TOTALFIXEDASSETS'.lower(), str(ws.cell(row, col).value).replace(' ', '').lower()) or \
                                                            re.search('CURRENTASSETS'.lower(), str(ws.cell(row, col).value).replace(' ', '').lower()):
                                                        negative_start = True
                                                    total_str = str(ws.cell(row, col).value).replace('=', '').replace('"', '')
                                                    col += 1
                                                    while ws.cell(row, col).value is None and col < ws.max_column:
                                                        col += 1
                                                    if re.search(sum_pattern_1, str(ws.cell(row, col).value)) or re.search(sum_pattern_2, str(ws.cell(row, col).value)):
                                                        result = get_result(str(ws.cell(row, col).value), value_array)
                                                        self.set_cell_style((ws.cell(row, col)))
                                                        elements = self.total_style(elements, total_str, result, colWidths, decimal_place)
                                                        value_array[row] = float(result)
                                                    else:
                                                        elements = self.normal_style(elements, total_str, colWidths)
                                                break
                                            elif re.search(acc_pattern_4, str(ws.cell(row, col).value)):  # "2000:2100,2102%%"
                                                empty_row = False
                                                acc_string = str(ws.cell(row, col).value)
                                                col += 1
                                                while ws.cell(row, col).value is None or str(ws.cell(row, col).value) == 'D(ACCT)':
                                                    col += 1
                                                if re.search(sum_pattern_3, str(ws.cell(row, col).value)):
                                                    val_1 = value_array[row - 1]
                                                    val_2 = value_array[row - 2]
                                                    result = val_1 + val_2
                                                    value_array[row] = float(result)
                                                    str_sum_change, style_sum_change = wrap_text(result, decimal_place)
                                                    table_data = []
                                                    table_data.append(
                                                        ['', '', '', '', Paragraph(str_sum_change, styles[style_sum_change]), ''])
                                                    item_table = Table(table_data, colWidths=colWidths, rowHeights=rowHeights)
                                                    item_table.setStyle(TableStyle(
                                                        [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                         ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                         ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                         ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
                                                         ('LEFTPADDING', (0, 0), (-1, -1), 0),
                                                         ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                                                         ('LINEABOVE', (4, 0), (4, 0), 0.25, colors.black),
                                                         ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                                                         ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                                                         ]))
                                                    elements.append(item_table)
                                                else:
                                                    acc_str_array = acc_string.split(',')
                                                    acc_ids = []
                                                    for acc_str in acc_str_array:
                                                        if re.search('[A-Z]*[0-9]+[:][A-Z]*[0-9]+[Z]*', acc_str):
                                                            acc_range = acc_str
                                                            acc_range = acc_range.replace('Z', '9').replace('=', '').replace('"', '').strip().split(':')
                                                            f_code = int(re.sub('[A-Z]+', '', acc_range[0]))
                                                            l_code = int(re.sub('[A-Z]+', '', acc_range[1]))
                                                            code_list = []
                                                            while f_code <= l_code:
                                                                code_list.append(str(f_code))
                                                                f_code += 1
                                                            ids = account_item_list.filter(account_segment__in=code_list).order_by('id').values_list('id', flat=True)
                                                            acc_ids = acc_ids + list(ids)
                                                        elif re.search('[A-Z]*[0-9]+[%]*', acc_str):
                                                            f_code = acc_str.replace('%', '').replace('=', '').replace('"', '').strip()
                                                            f_code = re.sub('[A-Z]+', '', f_code)
                                                            ids = account_item_list.filter(account_segment=f_code).order_by('id').values_list('id', flat=True)
                                                            acc_ids = acc_ids + list(ids)

                                                    account_list = account_item_list.filter(id__in=acc_ids).order_by('account_segment', 'code')
                                                    if filter_type == 'Segment':
                                                        account_list = account_list.filter(segment_code_id__in=segment_ids)
                                                    if re.search('ACCTDESC'.lower(), str(ws.cell(row, col).value).lower()):
                                                        accumulate = False
                                                    else:
                                                        accumulate = True
                                                    if re.search('FORTHEPERIOD'.lower(), str(ws.cell(row, col).value).replace(' ', '').lower()):
                                                        profit_loss = True
                                                    else:
                                                        profit_loss = False
                                                    name_str = str(ws.cell(row, col).value)

                                                    col += 1
                                                    while not re.search(sum_pattern_3, str(ws.cell(row, col).value)) and col < ws.max_column:
                                                        col += 1
                                                    sign = Decimal(1.0)
                                                    if re.search('\-', str(ws.cell(row, col).value)) and negative_start:
                                                        sign = Decimal(-1.0)
                                                    if self.report_type == dict(REPORT_TYPE)['Provisional'] and profit_loss:
                                                        sign = Decimal(1.0)
                                                    code_list = account_list.order_by('account_segment').values_list('account_segment', flat=True).distinct()
                                                    accumulate_total = 0
                                                    for cod in code_list:
                                                        total_account = 0
                                                        segments = account_list.filter(account_segment=cod)
                                                        for f_acc in segments:
                                                            item_account_q = accountHistory.filter(account_id=f_acc.id)
                                                            item_account = item_account_q.filter(
                                                                Q(source_currency_id=company.currency_id) |
                                                                Q(source_currency_id__isnull=True))
                                                            if item_account:
                                                                sum_debit = sum_credit = 0
                                                                for iAccount in item_account:
                                                                    if f_acc.balance_type == BALANCE_TYPE_DICT['Credit']:
                                                                        sum_credit += (iAccount.functional_end_balance * sign)
                                                                    if f_acc.balance_type == BALANCE_TYPE_DICT['Debit']:
                                                                        sum_debit += (iAccount.functional_end_balance * sign)
                                                                total_segment = sum_debit + sum_credit
                                                                if self.report_type == dict(REPORT_TYPE)['Provisional']:
                                                                    prov_posted = getProvisionalPostedAmount(company_id, int(array_data[0]), int(array_data[1]),
                                                                                                             f_acc.id, iAccount.source_currency_id)
                                                                    total_segment += Decimal(prov_posted)
                                                                total_account += total_segment
                                                                accumulate_total += total_segment
                                                                str_sum_change, style_sum_change = wrap_text(total_segment, decimal_place)
                                                                if not accumulate:  # One segment
                                                                    if filter_type == 'Segment':
                                                                        table_data = []
                                                                        table_data.append(
                                                                            ['', f_acc.code, str(f_acc.name), '',
                                                                             Paragraph(str_sum_change, styles[style_sum_change]), ''])
                                                                        item_table = Table(table_data, colWidths=colWidths, rowHeights=rowHeights)
                                                                        item_table.setStyle(TableStyle(
                                                                            [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                                             ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                                             ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                                             ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
                                                                             ('LEFTPADDING', (0, 0), (-1, -1), 0),
                                                                             ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                                                                             ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                                                                             ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                                                                             ]))
                                                                        elements.append(item_table)
                                                        if not accumulate:  # All segment
                                                            if filter_type == 'Account':
                                                                f_acc = account_item_list.filter(account_segment=str(cod)).first()
                                                                if f_acc:
                                                                    code = f_acc.code.split('-')[0]
                                                                    if company.use_segment:
                                                                        names = f_acc.name.split('-')
                                                                        for cd in seg_codes:
                                                                            if names[-1] in cd:
                                                                                names.pop()
                                                                        name = '-'.join(names)
                                                                    else:
                                                                        name = f_acc.name
                                                                    str_sum_change, style_sum_change = wrap_text(total_account, decimal_place)
                                                                    table_data = []
                                                                    table_data.append(
                                                                        ['', code, str(name), '',
                                                                         Paragraph(str_sum_change, styles[style_sum_change]), ''])
                                                                    item_table = Table(table_data, colWidths=colWidths, rowHeights=rowHeights)
                                                                    item_table.setStyle(TableStyle(
                                                                        [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                                         ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                                         ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                                         ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
                                                                         ('LEFTPADDING', (0, 0), (-1, -1), 0),
                                                                         ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                                                                         ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                                                                         ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                                                                         ]))
                                                                    elements.append(item_table)
                                                    if self.report_type == dict(REPORT_TYPE)['Provisional'] and profit_loss:
                                                        accumulate_total *= -1
                                                    value_array[row] = float(accumulate_total)
                                                    if accumulate:  # accumulate_pattern
                                                        if re.search(accumulate_pattern, name_str):
                                                            name = re.sub(accumulate_pattern, '', name_str).replace('=', '').replace('"', '').strip()
                                                            code = re.sub(name, '', name_str).replace('=', '').replace('"', '').replace(' ', '').strip()
                                                            span = ('SPAN', (1, 0), (1, 0))
                                                        else:
                                                            code = name_str.replace('=', '').replace('"', '').strip()
                                                            name = ''
                                                            span = ('SPAN', (1, 0), (2, 0))

                                                        str_sum_change, style_sum_change = wrap_text(accumulate_total, decimal_place)
                                                        table_data = []
                                                        table_data.append(
                                                            ['', code, str(name), '',
                                                             Paragraph(str_sum_change, styles[style_sum_change]), ''])
                                                        item_table = Table(table_data, colWidths=colWidths, rowHeights=rowHeights)
                                                        item_table.setStyle(TableStyle(
                                                            [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                             ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                             ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                             ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
                                                             ('LEFTPADDING', (0, 0), (-1, -1), 0),
                                                             ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                                                             ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                                                             ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                                                             span,
                                                             ]))
                                                        elements.append(item_table)
                                                break
                                            elif re.search(acc_pattern_1, str(ws.cell(row, col).value)):
                                                empty_row = False
                                                acc_range = str(ws.cell(row, col).value)
                                                acc_range = acc_range.replace('Z', '9').replace('=', '').replace('"', '').strip().split(':')
                                                f_code = int(re.sub('[A-Z]+', '', acc_range[0]))
                                                l_code = int(re.sub('[A-Z]+', '', acc_range[1]))
                                                code_list = []
                                                while f_code <= l_code:
                                                    code_list.append(str(f_code))
                                                    f_code += 1
                                                account_list = account_item_list.filter(account_segment__in=code_list).order_by('account_segment', 'code')
                                                if filter_type == 'Segment':
                                                    account_list = account_list.filter(segment_code_id__in=segment_ids)
                                                col += 1
                                                while ws.cell(row, col).value is None or str(ws.cell(row, col).value) == 'D(ACCT)':
                                                    col += 1
                                                if re.search('ACCTDESC'.lower(), str(ws.cell(row, col).value).lower()):
                                                    accumulate = False
                                                else:
                                                    accumulate = True
                                                if re.search('FORTHEPERIOD'.lower(), str(ws.cell(row, col).value).replace(' ', '').lower()):
                                                    profit_loss = True
                                                else:
                                                    profit_loss = False
                                                name_str = str(ws.cell(row, col).value)

                                                col += 1
                                                while not re.search(sum_pattern_3, str(ws.cell(row, col).value)) and col < ws.max_column:
                                                    col += 1
                                                sign = Decimal(1.0)
                                                if re.search('\-', str(ws.cell(row, col).value)) and negative_start:
                                                    sign = Decimal(-1.0)
                                                if self.report_type == dict(REPORT_TYPE)['Provisional'] and profit_loss:
                                                    sign = Decimal(1.0)
                                                code_list = account_list.order_by('account_segment').values_list('account_segment', flat=True).distinct()
                                                accumulate_total = 0
                                                for cod in code_list:
                                                    total_account = 0
                                                    segments = account_list.filter(account_segment=cod)
                                                    for f_acc in segments:
                                                        item_account_q = accountHistory.filter(account_id=f_acc.id)
                                                        item_account = item_account_q.filter(
                                                            Q(source_currency_id=company.currency_id) |
                                                            Q(source_currency_id__isnull=True))
                                                        if item_account:
                                                            sum_debit = sum_credit = 0
                                                            for iAccount in item_account:
                                                                if f_acc.balance_type == BALANCE_TYPE_DICT['Credit']:
                                                                    sum_credit += (iAccount.functional_end_balance * sign)
                                                                if f_acc.balance_type == BALANCE_TYPE_DICT['Debit']:
                                                                    sum_debit += (iAccount.functional_end_balance * sign)
                                                            total_segment = sum_debit + sum_credit
                                                            if self.report_type == dict(REPORT_TYPE)['Provisional']:
                                                                prov_posted = getProvisionalPostedAmount(company_id, int(array_data[0]), int(array_data[1]),
                                                                                                         f_acc.id, iAccount.source_currency_id)
                                                                total_segment += Decimal(prov_posted)
                                                            total_account += total_segment
                                                            accumulate_total += total_segment
                                                            str_sum_change, style_sum_change = wrap_text(total_segment, decimal_place)
                                                            if not accumulate:  # One segment
                                                                if filter_type == 'Segment':
                                                                    table_data = []
                                                                    table_data.append(
                                                                        ['', f_acc.code, str(f_acc.name), '',
                                                                         Paragraph(str_sum_change, styles[style_sum_change]), ''])
                                                                    item_table = Table(table_data, colWidths=colWidths, rowHeights=rowHeights)
                                                                    item_table.setStyle(TableStyle(
                                                                        [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                                         ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                                         ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                                         ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
                                                                         ('LEFTPADDING', (0, 0), (-1, -1), 0),
                                                                         ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                                                                         ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                                                                         ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                                                                         ]))
                                                                    elements.append(item_table)
                                                    if not accumulate:  # All segment
                                                        if filter_type == 'Account':
                                                            f_acc = account_item_list.filter(account_segment=str(cod)).first()
                                                            if f_acc:
                                                                code = f_acc.code.split('-')[0]
                                                                if company.use_segment:
                                                                    names = f_acc.name.split('-')
                                                                    for cd in seg_codes:
                                                                        if names[-1] in cd:
                                                                            names.pop()
                                                                    name = '-'.join(names)
                                                                else:
                                                                    name = f_acc.name
                                                                str_sum_change, style_sum_change = wrap_text(total_account, decimal_place)
                                                                table_data = []
                                                                table_data.append(['', code, str(name), '',
                                                                                   Paragraph(str_sum_change, styles[style_sum_change]), ''])
                                                                item_table = Table(table_data, colWidths=colWidths, rowHeights=rowHeights)
                                                                item_table.setStyle(TableStyle(
                                                                    [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                                     ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                                     ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                                     ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
                                                                     ('LEFTPADDING', (0, 0), (-1, -1), 0),
                                                                     ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                                                                     ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                                                                     ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                                                                     ]))
                                                                elements.append(item_table)
                                                if self.report_type == dict(REPORT_TYPE)['Provisional'] and profit_loss:
                                                    accumulate_total *= -1
                                                value_array[row] = float(accumulate_total)
                                                if accumulate:  # accumulate_pattern
                                                    if re.search(accumulate_pattern, name_str):
                                                        name = re.sub(accumulate_pattern, '', name_str).replace('=', '').replace('"', '').strip()
                                                        code = re.sub(name, '', name_str).replace('=', '').replace('"', '').replace(' ', '').strip()
                                                        span = ('SPAN', (1, 0), (1, 0))
                                                    else:
                                                        code = name_str.replace('=', '').replace('"', '').strip()
                                                        name = ''
                                                        span = ('SPAN', (1, 0), (2, 0))

                                                    str_sum_change, style_sum_change = wrap_text(accumulate_total, decimal_place)
                                                    table_data = []
                                                    table_data.append(
                                                        ['', code, str(name), '',
                                                         Paragraph(str_sum_change, styles[style_sum_change]), ''])
                                                    item_table = Table(table_data, colWidths=colWidths, rowHeights=rowHeights)
                                                    item_table.setStyle(TableStyle(
                                                        [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                         ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                         ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                         ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
                                                         ('LEFTPADDING', (0, 0), (-1, -1), 0),
                                                         ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                                                         ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                                                         ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                                                         span,
                                                         ]))
                                                    elements.append(item_table)
                                                break
                                            elif re.search(acc_pattern_3, str(ws.cell(row, col).value)):  # "2000%%,2100%%,2102%%"
                                                empty_row = False
                                                account_str = str(ws.cell(row, col).value)
                                                col += 1
                                                while ws.cell(row, col).value is None or str(ws.cell(row, col).value) == 'D(ACCT)':
                                                    col += 1
                                                if re.search(sum_pattern_3, str(ws.cell(row, col).value)):
                                                    val_1 = value_array[row - 1]
                                                    val_2 = value_array[row - 2]
                                                    result = val_1 + val_2
                                                    value_array[row] = float(result)
                                                    str_sum_change, style_sum_change = wrap_text(result, decimal_place)
                                                    table_data = []
                                                    table_data.append(['', '', '', '', Paragraph(str_sum_change, styles[style_sum_change]), ''])
                                                    item_table = Table(table_data, colWidths=colWidths, rowHeights=rowHeights)
                                                    item_table.setStyle(TableStyle(
                                                        [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                         ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                         ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                         ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
                                                         ('LEFTPADDING', (0, 0), (-1, -1), 0),
                                                         ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                                                         ('LINEABOVE', (4, 0), (4, 0), 0.25, colors.black),
                                                         ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                                                         ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                                                         ]))
                                                    elements.append(item_table)
                                                else:
                                                    name_str = str(ws.cell(row, col).value)
                                                    col += 1
                                                    while not re.search(sum_pattern_3, str(ws.cell(row, col).value)) and col < ws.max_column:
                                                        col += 1
                                                    sign = Decimal(1.0)
                                                    if re.search('\-', str(ws.cell(row, col).value)) and negative_start:
                                                        sign = Decimal(-1.0)
                                                    acc_list = account_str.replace('%', '').replace('=', '').replace('"', '').strip()
                                                    acc_list = re.sub('[A-Z]+', '', acc_list).split(',')
                                                    account_list = account_item_list.filter(account_segment__in=acc_list).order_by('account_segment', 'code')
                                                    if filter_type == 'Segment':
                                                        account_list = account_list.filter(segment_code_id__in=segment_ids)
                                                    accumulate_total = 0
                                                    for f_acc in account_list:
                                                        item_account_q = accountHistory.filter(account_id=f_acc.id)
                                                        item_account = item_account_q.filter(
                                                            Q(source_currency_id=company.currency_id) | Q(source_currency_id__isnull=True))

                                                        if item_account:
                                                            sum_debit = sum_credit = 0
                                                            for iAccount in item_account:
                                                                if f_acc.balance_type == BALANCE_TYPE_DICT['Credit']:
                                                                    sum_credit += (iAccount.functional_end_balance * sign)
                                                                if f_acc.balance_type == BALANCE_TYPE_DICT['Debit']:
                                                                    sum_debit += (iAccount.functional_end_balance * sign)
                                                            total_account = sum_debit + sum_credit
                                                            if self.report_type == dict(REPORT_TYPE)['Provisional']:
                                                                prov_posted = getProvisionalPostedAmount(company_id, int(array_data[0]), int(array_data[1]),
                                                                                                         f_acc.id, iAccount.source_currency_id)
                                                                total_account += Decimal(prov_posted)
                                                            accumulate_total += total_account

                                                    value_array[row] = float(accumulate_total)
                                                    if re.search(accumulate_pattern, name_str):
                                                        name = re.sub(accumulate_pattern, '', name_str).replace('=', '').replace('"', '').strip()
                                                        code = re.sub(name, '', name_str).replace('=', '').replace('"', '').replace(' ', '').strip()
                                                        span = ('SPAN', (1, 0), (1, 0))
                                                    else:
                                                        code = name_str.replace('=', '').replace('"', '').strip()
                                                        name = ''
                                                        span = ('SPAN', (1, 0), (2, 0))
                                                    str_sum_change, style_sum_change = wrap_text(accumulate_total, decimal_place)
                                                    table_data = []
                                                    table_data.append(
                                                        ['', code, str(name), '',
                                                         Paragraph(str_sum_change, styles[style_sum_change]), ''])
                                                    item_table = Table(table_data, colWidths=colWidths, rowHeights=rowHeights)
                                                    item_table.setStyle(TableStyle(
                                                        [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                         ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                         ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                         ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
                                                         ('LEFTPADDING', (0, 0), (-1, -1), 0),
                                                         ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                                                         ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                                                         ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                                                         span,
                                                         ]))
                                                    elements.append(item_table)
                                                break
                                            elif re.search(acc_pattern_2, str(ws.cell(row, col).value)):
                                                empty_row = False
                                                f_code = str(ws.cell(row, col).value).replace('%', '').replace('=', '').replace('"', '').strip()
                                                f_code = re.sub('[A-Z]+', '', f_code)
                                                col += 1
                                                while ws.cell(row, col).value is None or str(ws.cell(row, col).value) == 'D(ACCT)':
                                                    col += 1
                                                if re.search(sum_pattern_3, str(ws.cell(row, col).value)):
                                                    result = value_array[row - 1]
                                                    value_array[row] = float(result)
                                                    str_sum_change, style_sum_change = wrap_text(result, decimal_place)
                                                    table_data = []
                                                    table_data.append(
                                                        ['', '', '', '', Paragraph(str_sum_change, styles[style_sum_change]), ''])
                                                    item_table = Table(table_data, colWidths=colWidths, rowHeights=rowHeights)
                                                    item_table.setStyle(TableStyle(
                                                        [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                         ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                         ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                         ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
                                                         ('LEFTPADDING', (0, 0), (-1, -1), 0),
                                                         ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                                                         ('LINEABOVE', (4, 0), (4, 0), 0.25, colors.black),
                                                         ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                                                         ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                                                         ]))
                                                    elements.append(item_table)
                                                else:
                                                    name_str = str(ws.cell(row, col).value)

                                                    col += 1
                                                    while not re.search(sum_pattern_3, str(ws.cell(row, col).value)) and col < ws.max_column:
                                                        col += 1
                                                    sign = Decimal(1.0)
                                                    if re.search('\-', str(ws.cell(row, col).value)) and negative_start:
                                                        sign = Decimal(-1.0)
                                                    account_list = account_item_list.filter(account_segment=str(f_code))
                                                    if filter_type == 'Segment':
                                                        account_list = account_list.filter(segment_code_id__in=segment_ids)
                                                    segments_total = 0
                                                    if account_list:
                                                        for f_acc in account_list:
                                                            item_account_q = accountHistory.filter(account_id=f_acc.id)
                                                            item_account = item_account_q.filter(
                                                                Q(source_currency_id=company.currency_id) | Q(source_currency_id__isnull=True))

                                                            if item_account:
                                                                sum_debit = sum_credit = 0
                                                                for iAccount in item_account:
                                                                    if f_acc.balance_type == BALANCE_TYPE_DICT['Credit']:
                                                                        sum_credit += (iAccount.functional_end_balance * sign)
                                                                    if f_acc.balance_type == BALANCE_TYPE_DICT['Debit']:
                                                                        sum_debit += (iAccount.functional_end_balance * sign)
                                                                total_account = (sum_debit + sum_credit)
                                                                if self.report_type == dict(REPORT_TYPE)['Provisional']:
                                                                    prov_posted = getProvisionalPostedAmount(company_id, int(array_data[0]), int(array_data[1]),
                                                                                                             f_acc.id, iAccount.source_currency_id)
                                                                    total_account += Decimal(prov_posted)
                                                                segments_total += total_account
                                                                if filter_type == 'Segment':
                                                                    if re.search(code_pattern_1, name_str) and re.search(name_pattern, name_str) or \
                                                                            re.search(code_pattern_2, name_str) and re.search(name_pattern, name_str):
                                                                        code = str(f_acc.code)
                                                                        name = str(f_acc.name)
                                                                    elif re.search(code_pattern_1, name_str) or re.search(code_pattern_2, name_str):
                                                                        code = str(f_acc.code)
                                                                        name = re.sub(code_pattern_1, '', name_str)
                                                                        name = re.sub(code_pattern_2, '', name)
                                                                        name = name.replace('&', '').replace('=', '').replace('"', '').strip()
                                                                    elif re.search(str(f_acc.account_segment), name_str):
                                                                        code = str(f_acc.code)
                                                                        name = re.sub(str(f_acc.account_segment), '', name_str)
                                                                        name = name.replace('&', '').replace('=', '').replace('"', '').strip()
                                                                        if re.search('ACCTDESC'.lower(), str(name).lower()) or re.search('NAME'.lower(), str(name).lower()):
                                                                            name = f_acc.name
                                                                    else:
                                                                        code = ''
                                                                        name = name_str.replace('&', '').replace('=', '').replace('"', '')

                                                                    str_sum_change, style_sum_change = wrap_text(total_account, decimal_place)
                                                                    table_data = []
                                                                    table_data.append(
                                                                        ['', code, str(name), '',
                                                                         Paragraph(str_sum_change, styles[style_sum_change]), ''])
                                                                    item_table = Table(table_data, colWidths=colWidths, rowHeights=rowHeights)
                                                                    item_table.setStyle(TableStyle(
                                                                        [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                                         ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                                         ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                                         ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
                                                                         ('LEFTPADDING', (0, 0), (-1, -1), 0),
                                                                         ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                                                                         ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                                                                         ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                                                                         ]))
                                                                    elements.append(item_table)
                                                        value_array[row] = float(segments_total)
                                                        if filter_type == 'Account':
                                                            s_account = account_item_list.filter(account_segment=str(f_code)).first()
                                                            if re.search(str(f_acc.account_segment), name_str):
                                                                code = f_acc.code.split('-')[0]
                                                                name = re.sub(str(f_acc.account_segment), '', name_str)
                                                                name = name.replace('&', '').replace('=', '').replace('"', '').strip()
                                                                if re.search('ACCTDESC'.lower(), str(name).lower()) or re.search('NAME'.lower(), str(name).lower()):
                                                                    if company.use_segment:
                                                                        names = f_acc.name.split('-')
                                                                        for cd in seg_codes:
                                                                            if names[-1] in cd:
                                                                                names.pop()
                                                                        name = '-'.join(names)
                                                                    else:
                                                                        name = f_acc.name
                                                            else:
                                                                code = s_account.code.split('-')[0]
                                                                if re.search('ACCTDESC'.lower(), name_str.lower()) or re.search('NAME'.lower(), name_str.lower()):
                                                                    if company.use_segment:
                                                                        names = s_account.name.split('-')
                                                                        for cd in seg_codes:
                                                                            if names[-1] in cd:
                                                                                names.pop()
                                                                        name = '-'.join(names)
                                                                    else:
                                                                        name = s_account.name
                                                                else:
                                                                    name_str = re.sub(code_pattern_1, '', name_str)
                                                                    name = name_str.replace('&', '').replace('=', '').replace('"', '').strip()
                                                            str_sum_change, style_sum_change = wrap_text(segments_total, decimal_place)
                                                            table_data = []
                                                            table_data.append(
                                                                ['', code, str(name), '',
                                                                 Paragraph(str_sum_change, styles[style_sum_change]), ''])
                                                            item_table = Table(table_data, colWidths=colWidths, rowHeights=rowHeights)
                                                            item_table.setStyle(TableStyle(
                                                                [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                                 ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                                 ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                                 ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
                                                                 ('LEFTPADDING', (0, 0), (-1, -1), 0),
                                                                 ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                                                                 ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                                                                 ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                                                                 ]))
                                                            elements.append(item_table)
                                                break
                                        else:
                                            empty_row = False
                            else:
                                if empty_row:
                                    elements = self.normal_style(elements, '', colWidths)
                except Exception as e:
                    print(e)
                    table_data = []
                    table_data.append(['', '', ''])
                    table_data.append(['', 'ERROR! ON FILE PROCESSING', ''])
                    table_data.append(['', 'PLEASE MAKE SURE IT IS CORRECT TEMPLATE.', ''])

                    item_table = Table(table_data, colWidths=[100, 310, 100])
                    item_table.setStyle(TableStyle(
                        [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                         ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                         ('FONT', (0, 0), (-1, -1), s.REPORT_FONT_BOLD),
                         ('FONTSIZE', (0, 0), (-1, -1), 20),
                         ('SPAN', (1, 1), (2, 1)),
                         ('SPAN', (1, 2), (2, 2)),
                         ]))
                    elements.append(item_table)

            except Exception as e:
                print(e)
                table_data = []
                table_data.append(['', '', ''])
                table_data.append(['', '', ''])
                table_data.append(['', 'ERROR! CAN NOT READ THE TEMPLATE.', ''])

                item_table = Table(table_data, colWidths=[100, 310, 100])
                item_table.setStyle(TableStyle(
                    [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                     ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                     ('FONT', (0, 0), (-1, -1), s.REPORT_FONT_BOLD),
                     ('FONTSIZE', (0, 0), (-1, -1), 20),
                     ('SPAN', (1, 2), (-1, -1)),
                     ]))
                elements.append(item_table)

        else:  # No template
            table_data = []
            table_data.append(['', '', ''])
            table_data.append(['', '', ''])
            table_data.append(['', 'ERROR! THERE IS NO TEMPLATE TO PROCESS.', ''])

            item_table = Table(table_data, colWidths=[100, 310, 100])
            item_table.setStyle(TableStyle(
                [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                 ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                 ('FONT', (0, 0), (-1, -1), s.REPORT_FONT_BOLD),
                 ('FONTSIZE', (0, 0), (-1, -1), 20),
                 ('SPAN', (1, 2), (-1, -1)),
                 ]))
            elements.append(item_table)

        doc.build(elements,
                  onFirstPage=partial(self._header_footer, company_id=company_id, issue_from=issue_from,
                                      issue_to=issue_to
                                      ),
                  onLaterPages=partial(self._header_footer, company_id=company_id, issue_from=issue_from,
                                       issue_to=issue_to),
                  canvasmaker=partial(NumberedPage, adjusted_height=s.REPORT_PAGE_NO_HEIGHT, adjusted_width=255))
        # Get the value of the BytesIO buffer and write it to the response.
        pdf = buffer.getvalue()
        buffer.close()

        return pdf

    def normal_style(self, elements, text, colWidths):
        table_data = []
        table_data.append(['', text.replace('=', '').replace('"', ''), '', '', '', ''])
        item_table = Table(table_data, colWidths=colWidths)
        item_table.setStyle(TableStyle(
            [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
             ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
             ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
             ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
             ('LEFTPADDING', (0, 0), (-1, -1), 0),
             ('RIGHTPADDING', (0, 0), (-1, -1), 0),
             ('SPAN', (1, 0), (2, 0)),
             ('ALIGN', (1, 0), (2, 0), 'LEFT'),
             ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
             ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
             ]))
        elements.append(item_table)
        return elements

    def total_style(self, elements, text, total, colWidths, decimal_place):
        str_sum_change, style_sum_change = wrap_text(total, decimal_place)
        table_data = []
        table_data.append(
            ['', text.replace('=', '').replace('"', ''),
                '', '', Paragraph(str_sum_change, styles[style_sum_change]), ''])
        item_table = Table(table_data, colWidths=colWidths)

        above_border = 1
        above_color = colors.transparent
        below_border = 1
        below_color = colors.transparent
        if self.top_style == 'double':
            above_border = 2
            above_color = colors.black
        if self.top_style == 'thin':
            above_color = colors.black
        if self.bottom_style == 'double':
            below_border = 2
            below_color = colors.black
        if self.bottom_style == 'thin':
            below_color = colors.black
        item_table.setStyle(TableStyle(
            [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
             ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
             ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
             ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
             ('LEFTPADDING', (0, 0), (-1, -1), 0),
             ('RIGHTPADDING', (0, 0), (-1, -1), 0),
             ('SPAN', (1, 0), (2, 0)),
             ('ALIGN', (1, 0), (2, 0), 'LEFT'),
             ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
             ('LINEABOVE', (4, 0), (4, 0), 0.25, above_color, 0, None, None, above_border, 2),
             ('LINEBELOW', (4, 0), (4, 0), 0.25, below_color, 0, None, None, below_border, 2),
             ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
             ]))
        elements.append(item_table)
        return elements

    def set_cell_style(self, cell):
        if cell.has_style:
            if cell.border.bottom.border_style == 'double' or \
                cell.border.bottom.border_style == 'thin':
                self.bottom_style = cell.border.bottom.border_style
            else:
                self.bottom_style = ''
            if cell.border.top.border_style == 'double' or \
                cell.border.top.border_style == 'thin':
                self.top_style = cell.border.top.border_style
            else:
                self.top_style = ''
        else:
            self.top_style = ''
            self.bottom_style = ''

    def createCustomAccGroupList(self, company_id):
        rpt_acc_grp_list = []
        rpt_acct_grp = ReportGroup.objects.filter(company_id=company_id, is_hidden=False,
                                                  report_template_type=REPORT_TEMPLATE_TYPES_DICT['Balance Sheet'])
        for rpt_acc_grp in rpt_acct_grp:
            rpt_acc_grp_obj = {'acc1_id': rpt_acc_grp.account_from.id,
                               'acc2_id': rpt_acc_grp.account_to.id if rpt_acc_grp.account_to else rpt_acc_grp.account_from.id,
                               'acc_code': rpt_acc_grp.account_code_text,
                               'acc_name': rpt_acc_grp.name}
            rpt_acc_grp_list.append(rpt_acc_grp_obj)
        return rpt_acc_grp_list

    def getCustomAccGroup(self, rpt_acc_grp_list, mAccount):
        custom_acc = {"custom_acc_code": None, "custom_acc_name": None}
        for custom_acct in rpt_acc_grp_list:
            if mAccount.id >= custom_acct['acc1_id'] and mAccount.id <= custom_acct['acc2_id']:
                custom_acc['custom_acc_code'] = custom_acct['acc_code']
                custom_acc['custom_acc_name'] = str(custom_acct['acc_name'])
                break
            else:
                custom_acc['custom_acc_code'] = mAccount.code
                custom_acc['custom_acc_name'] = str(mAccount.name)
        return custom_acc
