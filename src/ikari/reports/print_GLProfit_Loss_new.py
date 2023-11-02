import calendar
import datetime
import os
import re
from openpyxl import load_workbook
from decimal import Decimal
from functools import partial
from django.db.models import Q
from django.conf import settings as s
from django.core.files.base import ContentFile
from reportlab.lib import colors
from reportlab.lib.enums import TA_RIGHT, TA_LEFT
from reportlab.lib.pagesizes import A4, letter, portrait
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from accounting.models import FiscalCalendar
from accounts.models import Account, AccountHistory, ReportGroup
from currencies.models import ExchangeRate
from companies.models import Company, CostCenters
from reports.numbered_page import NumberedPage
from transactions.models import Transaction
from reports.print_GLBalance_Sheet_XLS_new import get_result
from utilities.common import round_number
from utilities.constants import STATUS_TYPE_DICT, REPORT_TYPE, MONTHS_STR_DICT
from django.contrib.humanize.templatetags.humanize import intcomma

rowHeights = 14
colWidths = [10, 80, 240, 10, 80, 10, 80, 10]
styles = getSampleStyleSheet()
styles.add(ParagraphStyle(name='RightAlignBold', fontName=s.REPORT_FONT_BOLD, alignment=TA_RIGHT, fontSize=s.REPORT_FONT_SIZE))
styles.add(ParagraphStyle(name='LeftAlignBold', fontName=s.REPORT_FONT_BOLD, alignment=TA_LEFT, fontSize=s.REPORT_FONT_SIZE))
styles.add(ParagraphStyle(name='RightAlign', fontName=s.REPORT_FONT_BOLD, textColor=colors.red, alignment=TA_RIGHT, fontSize=s.REPORT_FONT_SIZE))
styles.add(ParagraphStyle(name='RightAlignBlack', fontName=s.REPORT_FONT, alignment=TA_RIGHT, fontSize=s.REPORT_FONT_SIZE))
styles.add(ParagraphStyle(name='RightAlignRed', fontName=s.REPORT_FONT, textColor=colors.black, alignment=TA_RIGHT, fontSize=s.REPORT_FONT_SIZE))


class Print_GLProfitLoss_new:
    def __init__(self, buffer, pagesize, report_type):
        self.buffer = buffer
        if pagesize == 'A4':
            self.pagesize = A4
        elif pagesize == 'Letter':
            self.pagesize = letter
        self.width, self.height = self.pagesize
        self.report_type = report_type
        self.pl_custom_acc_grp = None
        self.bottom_style = ''
        self.top_style = ''

    @staticmethod
    def _header_footer(canvas, doc, issue_from):
        pdfmetrics.registerFont(TTFont('Arial', os.path.join(s.BASE_DIR, "static/fonts/arial.ttf")))
        pdfmetrics.registerFont(TTFont('Arial-Bold', os.path.join(s.BASE_DIR, "static/fonts/arial-bold.ttf")))

        # Save the state of our canvas so we can draw on it
        canvas.saveState()
        # 1st row
        header_data = []
        row1_info1 = datetime.datetime.now().strftime('%d/%m/%Y %r')
        row1_info2 = ""
        header_data.append([row1_info1, row1_info2])

        header_table = Table(header_data, colWidths=[280, 330, 200])
        header_table.setStyle(TableStyle(
            [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
             ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
             ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
             ]))

        w, h = header_table.wrap(doc.width, doc.topMargin)
        header_table.drawOn(canvas, doc.leftMargin - 7, doc.height + doc.topMargin - h)
        # Release the canvas
        canvas.restoreState()

    def print_report(self, company_id, issue_from, filter_type, from_val, to_val):
        pdfmetrics.registerFont(TTFont('Arial', os.path.join(s.BASE_DIR, "static/fonts/arial.ttf")))
        pdfmetrics.registerFont(TTFont('Arial-Bold', os.path.join(s.BASE_DIR, "static/fonts/arial-bold.ttf")))

        buffer = self.buffer
        self.pagesize = portrait(A4)
        doc = SimpleDocTemplate(buffer, rightMargin=22, leftMargin=22, topMargin=50, bottomMargin=42, pagesize=portrait(A4))
        # Our container for 'Flowable' objects
        # Our container for 'Flowable' objects
        elements = []

        # Draw Content of PDF
        # design interface header
        array_data = str(issue_from).split('-')
        month = int(array_data[1])
        year = int(array_data[0])
        issue_from = datetime.date(int(array_data[0]), int(array_data[1]), 1)
        issue_to = datetime.date(int(array_data[0]), int(array_data[1]),
                                 calendar.monthrange(int(array_data[0]), int(array_data[1]))[1])
        company = Company.objects.get(pk=company_id)

        segment_ids = []
        segment_name = ''
        if filter_type == 'undefined' or filter_type == '' or filter_type == 'None':
            filter_type = 'Account'
        if filter_type == 'Segment':
            try:
                from_val = int(from_val)
                to_val = int(to_val)
                segments = CostCenters.objects.filter(id__range=(from_val, to_val))
                for seg in segments:
                    segment_name += ' - ' + seg.name
                segment_ids = segments.values_list('id', flat=True)
            except:
                pass

        fsc_calendar = FiscalCalendar.objects.filter(company_id=company_id, is_hidden=0, period=month, fiscal_year=year).last()
        if fsc_calendar:
            end_date = fsc_calendar.end_date.strftime('%d/%m/%Y')
        else:
            end_date = issue_to.strftime('%d/%m/%Y')

        accountHistory = AccountHistory.objects.select_related('account').filter(company_id=company.id, is_hidden=0,
                                                                                 period_month__exact=month,
                                                                                 period_year__exact=year)\
            .exclude(source_currency_id__isnull=True)

        decimal_place = "%.2f"
        if not company.currency.is_decimal:
            decimal_place = "%.0f"

        if company.pl_template:
            try:
                bytefile = company.pl_template.read()
                wb = load_workbook(ContentFile(bytefile))
                sheet_0 = wb.get_sheet_names()[0]
                ws = wb[sheet_0]

                hidden_rows = []
                for rowNum, rowDimension in ws.row_dimensions.items():
                    if rowDimension.hidden == True:
                        hidden_rows.append(rowNum)
                hidden_cols = []
                for colNum, colDimension in ws.column_dimensions.items():
                    if colDimension.hidden == True:
                        hidden_cols.append(colNum)

                acc_pattern_1 = '[=\"][A-Z]*[0-9]+[:][A-Z]*[0-9]+[Z]*[\"]'
                acc_pattern_2 = '[=\"][A-Z]*[0-9]+[%]*[\"]'
                sum_pattern_1 = re.compile(r'SUM')
                sum_pattern_2 = '[A-Z][0-9]+[+]*[-]*[A-Z]*[0-9]*'
                sum_pattern_3 = '[\\\\][-]*'
                sum_pattern_4 = '[0-9]+[:][0-9]+'
                change_array = [0] * (ws.max_row + 1)
                balance_array = [0] * (ws.max_row + 1)
                title_pattern = '[a-zA-Z]+[,]*[a-zA-Z]+[:]*[^0-9]+'
                start = False
                seg_codes = list(CostCenters.objects.filter(is_active=1, is_hidden=0, company_id=company_id).values_list('code', flat=True))
                account_item_list = Account.objects.filter(company_id=company.id, is_hidden=0).exclude(deactivate_period__lte=issue_from)
                account_item_list = account_item_list.exclude(Q(is_active=False) & Q(deactivate_date=None))
                try:
                    for row in range(1, ws.max_row + 1):
                        empty_row = True
                        if row not in hidden_rows:
                            for col in range(1, ws.max_column + 5):
                                if col not in hidden_cols:
                                    if ws.cell(row, col).value is not None:
                                        if start:
                                            if re.search(acc_pattern_1, str(ws.cell(row, col).value)):
                                                empty_row = False
                                                total_change = 0
                                                total_balance = 0
                                                acc_range = str(ws.cell(row, col).value)
                                                acc_range = acc_range.replace('Z', '9').replace('=', '').replace('"', '').strip().split(':')
                                                f_code = int(re.sub('[A-Z]+', '', acc_range[0]))
                                                l_code = int(re.sub('[A-Z]+', '', acc_range[1]))
                                                code_list = []
                                                while f_code <= l_code:
                                                    code_list.append(str(f_code))
                                                    f_code += 1
                                                account_list = account_item_list.filter(account_segment__in=code_list).order_by('account_segment', 'code')
                                                if filter_type == 'Segment':
                                                    account_list = account_list.filter(segment_code_id__in=segment_ids)
                                                code_list = account_list.order_by('account_segment').values_list('account_segment', flat=True).distinct()
                                                col += 1
                                                while not re.search(sum_pattern_3, str(ws.cell(row, col).value)) and col < ws.max_column:
                                                    col += 1
                                                sign = Decimal(1.0)
                                                if re.search('\-', str(ws.cell(row, col).value)):
                                                    sign = Decimal(-1.0)
                                                for cod in code_list:
                                                    acc_total_change = 0
                                                    acc_total_balance = 0
                                                    segments = account_list.filter(account_segment=cod)
                                                    for f_acc in segments:
                                                        item_account_q = accountHistory.filter(account_id=f_acc.id)
                                                        item_account = item_account_q.filter(
                                                            Q(source_currency_id=company.currency_id) |
                                                            Q(source_currency_id__isnull=True))
                                                        sum_change = sum_balance = 0
                                                        if item_account:
                                                            prov_posted_amount_per_account = 0
                                                            for iAccount in item_account:
                                                                if self.report_type == dict(REPORT_TYPE)['Provisional']:
                                                                    prov_posted_amount_per_account = getProvisionalPostedAmount(company.id, year, month,
                                                                                                                                f_acc.id, iAccount.source_currency_id)
                                                                sum_change += (iAccount.functional_net_change * sign) + Decimal(prov_posted_amount_per_account)
                                                                sum_balance += (iAccount.functional_end_balance * sign) + \
                                                                    Decimal(prov_posted_amount_per_account)
                                                                acc_total_change += (iAccount.functional_net_change * sign) + \
                                                                    Decimal(prov_posted_amount_per_account)
                                                                acc_total_balance += (iAccount.functional_end_balance * sign) + \
                                                                    Decimal(prov_posted_amount_per_account)
                                                                total_change += (iAccount.functional_net_change * sign) + \
                                                                    Decimal(prov_posted_amount_per_account)
                                                                total_balance += (iAccount.functional_end_balance * sign) + \
                                                                    Decimal(prov_posted_amount_per_account)
                                                                str_sum_change, style_sum_change = wrap_text(sum_change, decimal_place)
                                                                str_sum_balance, style_sum_balance = wrap_text(sum_balance, decimal_place)

                                                            if filter_type == 'Segment':
                                                                table_data = []
                                                                table_data.append(
                                                                    ['',
                                                                     f_acc.code,
                                                                     f_acc.name,
                                                                     '',
                                                                     Paragraph(str_sum_change if str_sum_change else str(sum_change), styles[style_sum_change]),
                                                                     '',
                                                                     Paragraph(str_sum_balance if str_sum_balance else str(
                                                                         sum_balance), styles[style_sum_balance]),
                                                                     ''])
                                                                item_table = Table(table_data, colWidths=colWidths, rowHeights=rowHeights)
                                                                item_table.setStyle(TableStyle(
                                                                    [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                                     ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                                     ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                                     ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
                                                                     ('LEFTPADDING', (0, 0), (-1, -1), 0),
                                                                     ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                                                                     ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                                                                     ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                                                                     ]))
                                                                elements.append(item_table)
                                                    if filter_type == 'Account':
                                                        f_acc = account_item_list.filter(account_segment=str(cod)).first()
                                                        if f_acc:
                                                            str_sum_change, style_sum_change = wrap_text(acc_total_change, decimal_place)
                                                            str_sum_balance, style_sum_balance = wrap_text(acc_total_balance, decimal_place)
                                                            code = f_acc.code.split('-')[0]
                                                            if company.use_segment:
                                                                names = f_acc.name.split('-')
                                                                if len(names) > 1:
                                                                    for cd in seg_codes:
                                                                        if names[-1] in cd:
                                                                            names.pop()
                                                                    name = '-'.join(names)
                                                                else:
                                                                    name = f_acc.name
                                                            else:
                                                                name = f_acc.name
                                                            table_data = []
                                                            table_data.append(
                                                                ['',
                                                                    code,
                                                                    name,
                                                                    '',
                                                                    Paragraph(str_sum_change if str_sum_change else str(
                                                                        acc_total_change), styles[style_sum_change]),
                                                                    '',
                                                                    Paragraph(str_sum_balance if str_sum_balance else str(
                                                                        acc_total_balance), styles[style_sum_balance]),
                                                                    ''])
                                                            item_table = Table(table_data, colWidths=colWidths, rowHeights=rowHeights)
                                                            item_table.setStyle(TableStyle(
                                                                [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                                 ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                                 ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                                 ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
                                                                 ('LEFTPADDING', (0, 0), (-1, -1), 0),
                                                                 ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                                                                 ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                                                                 ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                                                                 ]))
                                                            elements.append(item_table)
                                                change_array[row] = float(total_change)
                                                balance_array[row] = float(total_balance)
                                                break
                                            elif re.search(acc_pattern_2, str(ws.cell(row, col).value)):
                                                empty_row = False
                                                total_change = 0
                                                total_balance = 0
                                                f_code = str(ws.cell(row, col).value).replace('%', '').replace('=', '').replace('"', '').strip()
                                                f_code = re.sub('[A-Z]+', '', f_code)
                                                acc_list = account_item_list.filter(account_segment=str(f_code))
                                                if filter_type == 'Segment':
                                                    acc_list = acc_list.filter(segment_code_id__in=segment_ids)
                                                acc_total_change = 0
                                                acc_total_balance = 0
                                                col += 1
                                                while ws.cell(row, col).value is None or str(ws.cell(row, col).value) == 'D(ACCT)':
                                                    col += 1
                                                name_str = str(ws.cell(row, col).value)

                                                col += 1
                                                while not re.search(sum_pattern_3, str(ws.cell(row, col).value)) and col < ws.max_column:
                                                    col += 1
                                                sign = Decimal(1.0)
                                                if re.search('\-', str(ws.cell(row, col).value)):
                                                    sign = Decimal(-1.0)
                                                for f_acc in acc_list:
                                                    item_account_q = accountHistory.filter(account_id=f_acc.id)
                                                    item_account = item_account_q.filter(
                                                        Q(source_currency_id=company.currency_id) |
                                                        Q(source_currency_id__isnull=True))
                                                    sum_change = sum_balance = 0
                                                    if item_account:
                                                        prov_posted_amount_per_account = 0
                                                        for iAccount in item_account:
                                                            if self.report_type == dict(REPORT_TYPE)['Provisional']:
                                                                prov_posted_amount_per_account = getProvisionalPostedAmount(company.id, year, month,
                                                                                                                            f_acc.id, iAccount.source_currency_id)
                                                            sum_change += (iAccount.functional_net_change * sign) + Decimal(prov_posted_amount_per_account)
                                                            sum_balance += (iAccount.functional_end_balance * sign) + Decimal(prov_posted_amount_per_account)
                                                            acc_total_change += (iAccount.functional_net_change * sign) + \
                                                                Decimal(prov_posted_amount_per_account)
                                                            acc_total_balance += (iAccount.functional_end_balance * sign) + \
                                                                Decimal(prov_posted_amount_per_account)
                                                            total_change += (iAccount.functional_net_change * sign) + Decimal(prov_posted_amount_per_account)
                                                            total_balance += (iAccount.functional_end_balance * sign) + Decimal(prov_posted_amount_per_account)
                                                            str_sum_change, style_sum_change = wrap_text(sum_change, decimal_place)
                                                            str_sum_balance, style_sum_balance = wrap_text(sum_balance, decimal_place)

                                                        if filter_type == 'Segment':
                                                            if re.search('ACCTDESC"\)$', name_str):
                                                                code = str(f_acc.code)
                                                                name = str(f_acc.name)
                                                            elif re.search(str(f_acc.account_segment), name_str):
                                                                code = str(f_acc.code)
                                                                name = re.sub(str(f_acc.account_segment), '', name_str)
                                                                name = name.replace('&', '').replace('=', '').replace('"', '').strip()
                                                            else:
                                                                code = str(f_acc.code)
                                                                name = str(f_acc.name)

                                                            table_data = []
                                                            table_data.append(
                                                                ['',
                                                                    code,
                                                                    name,
                                                                    '',
                                                                    Paragraph(str_sum_change if str_sum_change else str(sum_change), styles[style_sum_change]),
                                                                    '',
                                                                    Paragraph(str_sum_balance if str_sum_balance else str(
                                                                        sum_balance), styles[style_sum_balance]),
                                                                    ''])
                                                            item_table = Table(table_data, colWidths=colWidths, rowHeights=rowHeights)
                                                            item_table.setStyle(TableStyle(
                                                                [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                                 ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                                 ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                                 ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
                                                                 ('LEFTPADDING', (0, 0), (-1, -1), 0),
                                                                 ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                                                                 ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                                                                 ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                                                                 ]))
                                                            elements.append(item_table)
                                                change_array[row] = float(total_change)
                                                balance_array[row] = float(total_balance)
                                                if filter_type == 'Account':
                                                    f_acc = account_item_list.filter(account_segment=str(f_code)).first()
                                                    if f_acc:
                                                        str_sum_change, style_sum_change = wrap_text(acc_total_change, decimal_place)
                                                        str_sum_balance, style_sum_balance = wrap_text(acc_total_balance, decimal_place)
                                                        code = f_acc.code.split('-')[0]
                                                        if company.use_segment:
                                                            names = f_acc.name.split('-')
                                                            if len(names) > 1:
                                                                for cd in seg_codes:
                                                                    if names[-1] in cd:
                                                                        names.pop()
                                                                name = '-'.join(names)
                                                            else:
                                                                name = f_acc.name
                                                        else:
                                                            name = f_acc.name
                                                        table_data = []
                                                        table_data.append([
                                                            '', code, name, '',
                                                            Paragraph(str_sum_change if str_sum_change else str(acc_total_change), styles[style_sum_change]),
                                                            '',
                                                            Paragraph(str_sum_balance if str_sum_balance else str(acc_total_balance), styles[style_sum_balance]), ''])
                                                        item_table = Table(table_data, colWidths=colWidths, rowHeights=rowHeights)
                                                        item_table.setStyle(TableStyle(
                                                            [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                             ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                             ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                             ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
                                                             ('LEFTPADDING', (0, 0), (-1, -1), 0),
                                                             ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                                                             ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                                                             ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                                                             ]))
                                                        elements.append(item_table)
                                                break
                                            elif re.search(title_pattern, str(ws.cell(row, col).value).replace(' ', '').lower()) or \
                                                    re.search(sum_pattern_1, str(ws.cell(row, col).value)) or re.search(sum_pattern_2, str(ws.cell(row, col).value)):
                                                empty_row = False
                                                if re.search(sum_pattern_1, str(ws.cell(row, col).value)) or re.search(sum_pattern_2, str(ws.cell(row, col).value)):
                                                    change = get_result(str(ws.cell(row, col).value), change_array)
                                                    col += 1
                                                    while ws.cell(row, col).value is None and col < ws.max_column:
                                                        col += 1
                                                    balance = get_result(str(ws.cell(row, col).value), balance_array)
                                                    self.set_cell_style((ws.cell(row, col)))

                                                    elements = self.total_style(elements, '', change, balance, decimal_place)

                                                    change_array[row] = float(change)
                                                    balance_array[row] = float(balance)
                                                else:
                                                    total_str = str(ws.cell(row, col).value).replace('=', '').replace('"', '')
                                                    col += 1
                                                    while ws.cell(row, col).value is None and col < ws.max_column:
                                                        col += 1
                                                    if col < ws.max_column:
                                                        if re.search(sum_pattern_1, str(ws.cell(row, col).value)) or re.search(sum_pattern_2, str(ws.cell(row, col).value)):
                                                            change = get_result(str(ws.cell(row, col).value), change_array)
                                                            col += 1
                                                            while ws.cell(row, col).value is None and col < ws.max_column:
                                                                col += 1
                                                            balance = get_result(str(ws.cell(row, col).value), balance_array)
                                                            self.set_cell_style((ws.cell(row, col)))

                                                            elements = self.total_style(elements, total_str, change, balance, decimal_place)

                                                            change_array[row] = float(change)
                                                            balance_array[row] = float(balance)

                                                    else:
                                                        elements = self.normal_style(elements, total_str)
                                                break
                                        else:
                                            empty_row = False
                                            if re.search('CONAME'.lower(), str(ws.cell(row, col).value).lower()):
                                                table_data = []
                                                table_data.append(['', str(company.name) + str(segment_name), ''])
                                                item_table = Table(table_data, colWidths=[100, 310, 100])
                                                item_table.setStyle(TableStyle(
                                                    [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                     ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                     ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                     ('FONTSIZE', (0, 0), (-1, -1), 12),
                                                     ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                                                     ]))
                                                elements.append(item_table)
                                                break
                                            elif re.search('PROFIT AND LOSS'.lower(), str(ws.cell(row, col).value).lower()):
                                                table_data = []
                                                table_data.append(['', 'TRADING AND PROFIT AND LOSS ACCOUNT', ''])
                                                item_table = Table(table_data, colWidths=[100, 310, 100])
                                                item_table.setStyle(TableStyle(
                                                    [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                     ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                     ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                     ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                                                     ]))
                                                elements.append(item_table)
                                                break
                                            elif re.search('FOR THE MONTH'.lower(), str(ws.cell(row, col).value).lower()):
                                                table_data = []
                                                table_data.append(['', "FOR THE MONTH OF " + end_date, ''])
                                                item_table = Table(table_data, colWidths=[100, 310, 100])
                                                item_table.setStyle(TableStyle(
                                                    [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                     ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                     ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                     ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                                                     ]))
                                                elements.append(item_table)
                                                break
                                            elif re.search(str(company.currency.code).lower(), str(ws.cell(row, col).value).lower()):
                                                start = True
                                                table_data = []
                                                table_data.append(['', '', 'CURRENT', '', 'YEAR TO', ''])
                                                table_data.append(['', '', 'MONTH', '', 'DATE', ''])
                                                table_data.append(['', '', company.currency.code, '', company.currency.code, ''])
                                                item_table = Table(table_data, colWidths=[10, 350, 70, 10, 70, 10], rowHeights=rowHeights)
                                                item_table.setStyle(TableStyle(
                                                    [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                        ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                                                        ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                                                        ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
                                                        ('LINEBELOW', (2, -1), (4, -1), 0.25, colors.black),
                                                        ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
                                                        ('LEFTPADDING', (0, 0), (-1, -1), 0),
                                                        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                                                        ('VALIGN', (0, 0), (-1, -1), 'TOP')
                                                     ]))
                                                elements.append(item_table)
                                                break
                            if empty_row:
                                elements = self.normal_style(elements, '')
                except Exception as e:
                    print(e)
                    table_data = []
                    table_data.append(['', '', ''])
                    table_data.append(['', 'ERROR! ON FILE PROCESSING', ''])
                    table_data.append(['', 'PLEASE MAKE SURE IT IS CORRECT TEMPLATE.', ''])

                    item_table = Table(table_data, colWidths=[100, 310, 100])
                    item_table.setStyle(TableStyle(
                        [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                         ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                         ('FONT', (0, 0), (-1, -1), s.REPORT_FONT_BOLD),
                         ('FONTSIZE', (0, 0), (-1, -1), 20),
                         ('SPAN', (1, 1), (2, 1)),
                         ('SPAN', (1, 2), (2, 2)),
                         ]))
                    elements.append(item_table)

            except:
                table_data = []
                table_data.append(['', '', ''])
                table_data.append(['', '', ''])
                table_data.append(['', 'ERROR! CAN NOT READ THE TEMPLATE.', ''])

                item_table = Table(table_data, colWidths=[100, 310, 100])
                item_table.setStyle(TableStyle(
                    [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                     ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                     ('FONT', (0, 0), (-1, -1), s.REPORT_FONT_BOLD),
                     ('FONTSIZE', (0, 0), (-1, -1), 20),
                     ('SPAN', (1, 2), (-1, -1)),
                     ]))
                elements.append(item_table)

        else:  # No template
            table_data = []
            table_data.append(['', '', ''])
            table_data.append(['', '', ''])
            table_data.append(['', 'ERROR! THERE IS NO TEMPLATE TO PROCESS.', ''])

            item_table = Table(table_data, colWidths=[100, 310, 100])
            item_table.setStyle(TableStyle(
                [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                 ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                 ('FONT', (0, 0), (-1, -1), s.REPORT_FONT_BOLD),
                 ('FONTSIZE', (0, 0), (-1, -1), 20),
                 ('SPAN', (1, 2), (-1, -1)),
                 ]))
            elements.append(item_table)
        # end process data
        # else:  # if there's no order in the selected month
        if elements.__len__() == 0:
            table_data = [['', '', '', '', '', '', '', '', '']]
            table_body = Table(table_data, colWidths=[60, 90, 110, 105, 75,
                                                      60, 85, 90, 130])
            table_body.setStyle(TableStyle(
                [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                 ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent)
                 ]))
            elements.append(table_body)

        doc.build(elements,
                  onFirstPage=partial(self._header_footer, issue_from=issue_from
                                      ),
                  onLaterPages=partial(self._header_footer, issue_from=issue_from),
                  canvasmaker=partial(NumberedPage, adjusted_height=140))
        # Get the value of the BytesIO buffer and write it to the response.
        pdf = buffer.getvalue()
        buffer.close()

        return pdf

    def normal_style(self, elements, text):
        table_data = []
        table_data.append(
            ['', text.replace('=', '').replace('"', ''), '', '',
                '', '', '', ''])
        item_table = Table(table_data, colWidths=colWidths)
        item_table.setStyle(TableStyle(
            [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
             ]))
        elements.append(item_table)
        return elements

    def set_cell_style(self, cell):
        if cell.has_style:
            if cell.border.bottom.border_style == 'double' or \
                cell.border.bottom.border_style == 'thin':
                self.bottom_style = cell.border.bottom.border_style
            else:
                self.bottom_style = ''
            if cell.border.top.border_style == 'double' or \
                cell.border.top.border_style == 'thin':
                self.top_style = cell.border.top.border_style
            else:
                self.top_style = ''
        else:
            self.top_style = ''
            self.bottom_style = ''
    
    def total_style(self, elements, text, total_change, total_balance, decimal_place):
        str_total_change, style_total_change = wrap_text(total_change, decimal_place)
        str_total_balance, style_total_balance = wrap_text(total_balance, decimal_place)
        
        above_border = 1
        above_color = colors.transparent
        below_border = 1
        below_color = colors.transparent
        if self.top_style == 'double':
            above_border = 2
            above_color = colors.black
        if self.top_style == 'thin':
            above_color = colors.black
        if self.bottom_style == 'double':
            below_border = 2
            below_color = colors.black
        if self.bottom_style == 'thin':
            below_color = colors.black
        table_data = []
        table_data.append(
            ['', text.replace('=', '').replace('"', ''), '', '',
                Paragraph(str_total_change, styles[style_total_change]),
                '',
                Paragraph(str_total_balance, styles[style_total_balance]), ''])
        item_table = Table(table_data, colWidths=colWidths)
        item_table.setStyle(TableStyle(
                [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.transparent),
                    ('BOX', (0, 0), (-1, -1), 0.25, colors.transparent),
                    ('FONT', (0, 0), (-1, -1), s.REPORT_FONT),
                    ('FONTSIZE', (0, 0), (-1, -1), s.REPORT_FONT_SIZE),
                    ('LEFTPADDING', (0, 0), (-1, -1), 0),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                    ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                    ('LINEABOVE', (4, 0), (6, 0), 0.25, above_color, 0, None, None, above_border, 2),
                    ('LINEBELOW', (4, 0), (6, 0), 0.25, below_color, 0, None, None, below_border, 2),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                 ]))
        elements.append(item_table)
        return elements


def getProvisionalPostedAmount(company_id, p_perd_year, p_perd_month, acc_id, curr):
    prov_posted_amount = 0
    provisional_transaction = Transaction.objects.filter(company_id=company_id, is_hidden=False,
                                                         journal__status=int(STATUS_TYPE_DICT['Prov. Posted']),
                                                         journal__perd_year__lte=p_perd_year,
                                                         journal__perd_month__lte=p_perd_month,
                                                         journal__is_hidden=False,
                                                         account_id=acc_id,
                                                         currency_id=curr)
    if provisional_transaction:
        for prov_trx in provisional_transaction:
            if prov_trx.account.balance_type != prov_trx.functional_balance_type:
                prov_posted_amount += (float(prov_trx.total_amount) * -1)
            else:
                prov_posted_amount += float(prov_trx.total_amount)
    return prov_posted_amount


def createCustomAccGroupList(company_id):
    rpt_acc_grp_list = []
    rpt_acct_grp = ReportGroup.objects.filter(company_id=company_id, is_hidden=False, report_template_type='0')
    for rpt_acc_grp in rpt_acct_grp:
        rpt_acc_grp_obj = {'acc1_id': rpt_acc_grp.account_from.id,
                           'acc2_id': rpt_acc_grp.account_to.id if rpt_acc_grp.account_to else rpt_acc_grp.account_from.id,
                           'acc_code': rpt_acc_grp.account_code_text,
                           'acc_name': rpt_acc_grp.name}
        rpt_acc_grp_list.append(rpt_acc_grp_obj)
    return rpt_acc_grp_list


def getCustomAccGroup(rpt_acc_grp_list, mAccount):
    is_4550_exist = False
    custom_acc = {"custom_acc_code": None, "custom_acc_name": None}
    for custom_acc in rpt_acc_grp_list:
        if mAccount.id >= custom_acc['acc1_id'] and mAccount.id <= custom_acc['acc2_id']:
            custom_acc['custom_acc_code'] = custom_acc['acc_code']
            custom_acc['custom_acc_name'] = str(custom_acc['acc_name'])
            if mAccount.code == '4550':
                is_4550_exist = True
            break
        else:
            custom_acc['custom_acc_code'] = mAccount.code
            custom_acc['custom_acc_name'] = str(mAccount.name)
    return custom_acc, is_4550_exist


def get_exchange_rate_header(issue_from, company_id):
    curr_month_exchange_rates = ExchangeRate.objects.filter(exchange_date=issue_from, company_id=company_id, is_hidden=0, flag='ACCOUNTING')
    array_data = str(issue_from).split('-')
    if int(array_data[1]) >= 1 and int(array_data[1]) <= 11:
        next_issue_from = datetime.date(int(array_data[0]), int(array_data[1]) + 1, 1)
    else:
        next_issue_from = datetime.date(int(array_data[0]) + 1, 1, 1)

    next_month_exchange_rates = ExchangeRate.objects.filter(exchange_date=next_issue_from, company_id=company_id, is_hidden=0, flag='ACCOUNTING')
    curr_usd_to_sgd = curr_month_exchange_rates.filter(from_currency__code='USD', to_currency__code='SGD').first()
    curr_yen_to_sgd = curr_month_exchange_rates.filter(from_currency__code='YEN', to_currency__code='SGD').first()
    next_usd_to_sgd = next_month_exchange_rates.filter(from_currency__code='USD', to_currency__code='SGD').first()
    next_yen_to_sgd = next_month_exchange_rates.filter(from_currency__code='YEN', to_currency__code='SGD').first()
    next_array_data = str(next_issue_from).split('-')

    if next_usd_to_sgd:
        frst_line = str(MONTHS_STR_DICT[next_array_data[1]]) + "'" + next_array_data[0] + ' - USD' + str(next_usd_to_sgd.rate)
    else:
        frst_line = ''
    if next_yen_to_sgd:
        scnd_line = '                  YEN' + str(next_yen_to_sgd.rate)
    else:
        scnd_line = ''
    if curr_usd_to_sgd:
        thrd_line = str(MONTHS_STR_DICT[array_data[1]]) + "'" + array_data[0] + ' - USD' + str(curr_usd_to_sgd.rate)
    else:
        thrd_line = ''
    if curr_yen_to_sgd:
        frth_line = '                  YEN' + str(curr_yen_to_sgd.rate)
    else:
        frth_line = ''

    return frst_line, scnd_line, thrd_line, frth_line


def wrap_text(value, decimal_place="%.2f"):
    str_value = intcomma(decimal_place % round_number(value)).replace("-", "")
    if value < 0:
        style = 'RightAlignRed'
        str_value = '(' + str_value + ')'
    else:
        style = 'RightAlignBlack'

    return str_value, style
