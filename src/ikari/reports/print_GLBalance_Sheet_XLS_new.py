import calendar
import datetime
import re
from openpyxl import load_workbook
from decimal import Decimal
import xlsxwriter
from django.db.models import Q
from django.core.files.base import ContentFile
from accounting.models import FiscalCalendar
from accounts.models import Account, AccountHistory, ReportGroup
from companies.models import Company, CostCenters
from transactions.models import Transaction
from utilities.common import round_number
from utilities.constants import STATUS_TYPE_DICT, BALANCE_TYPE_DICT, REPORT_TYPE
from django.contrib.humanize.templatetags.humanize import intcomma

filter_pattern = '[a-zA-Z]+'
acc_pattern_1 = '[=\"][A-Z]*[0-9]+[:][A-Z]*[0-9]+[Z]*[\"]'
acc_pattern_2 = '[=\"][A-Z]*[0-9]+[%]*[\"]'
acc_pattern_3 = '[=\"]*[A-Z]*[0-9]+[%]*[,][A-Z]*[0-9]+[%]*[\"]*'
acc_pattern_4 = '[=\"]*[A-Z]*[0-9]+[:][A-Z]*[0-9]+[Z]*[,][A-Z]*[0-9]+[%]*[\"]*'
accumulate_pattern = '[0-9]+[-][0-9]+'
sum_pattern_1 = re.compile(r'SUM')
sum_pattern_2 = '[G-H][0-9]+[+]*[-]*[G-H]*[0-9]*'
sum_pattern_3 = '[\\\\][-]*'
sum_pattern_4 = '[0-9]+[:][0-9]+'
code_pattern_1 = '[_xll\.]*FRACCT\(\"ACCTID\"\)'
code_pattern_2 = '[_xll\.]*FRACCT\(\"ACCTFMTTD\"\)'
name_pattern = '[_xll\.]*FRACCT\(\"ACCTDESC\"\)'


class Print_GLBalanceSheet_XLS_new:
    def __init__(self, buffer):
        self.buffer = buffer
        self.bottom_style = ''
        self.top_style = ''

    def createCustomAccGroupList(self, company_id):
        rpt_acc_grp_list = []
        rpt_acct_grp = ReportGroup.objects.filter(company_id=company_id, is_hidden=False, report_template_type='1')
        for rpt_acc_grp in rpt_acct_grp:
            rpt_acc_grp_obj = {'acc1_id': rpt_acc_grp.account_from.id,
                               'acc2_id': rpt_acc_grp.account_to.id if rpt_acc_grp.account_to else rpt_acc_grp.account_from.id,
                               'acc_code': rpt_acc_grp.account_code_text,
                               'acc_name': rpt_acc_grp.name}
            rpt_acc_grp_list.append(rpt_acc_grp_obj)
        return rpt_acc_grp_list

    def getCustomAccGroup(self, rpt_acc_grp_list, mAccount):
        custom_acc = {"custom_acc_code": None, "custom_acc_name": None}
        for custom_acc in rpt_acc_grp_list:
            if mAccount.id >= custom_acc['acc1_id'] and mAccount.id <= custom_acc['acc2_id']:
                custom_acc['custom_acc_code'] = custom_acc['acc_code']
                custom_acc['custom_acc_name'] = str(custom_acc['acc_name'])
                break
            else:
                custom_acc['custom_acc_code'] = mAccount.code
                custom_acc['custom_acc_name'] = str(mAccount.name)
        return custom_acc
    
    def set_cell_style(self, cell):
        if cell.has_style:
            if cell.border.bottom.border_style == 'double' or \
                cell.border.bottom.border_style == 'thin':
                self.bottom_style = cell.border.bottom.border_style
            else:
                self.bottom_style = ''
            if cell.border.top.border_style == 'double' or \
                cell.border.top.border_style == 'thin':
                self.top_style = cell.border.top.border_style
            else:
                self.top_style = ''
        else:
            self.top_style = ''
            self.bottom_style = ''

    def WriteToExcel(self, company_id, issue_date, report_type, filter_type, from_val, to_val):
        company = Company.objects.get(pk=company_id)
        output = self.buffer
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet("Balance Sheet")
        printing_row = 0
        printing_col = 0
        merge_format = workbook.add_format({
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
        })
        merge_str = workbook.add_format({
            'valign': 'vcenter',
        })
        right_line = workbook.add_format({
            'align': 'right'
        })
        center_line = workbook.add_format({
            'align': 'center',
            'bottom': 1
        })
        border_top_bot = workbook.add_format({
            'align': 'right',
            'bottom': 1,
            'top': 1
        })
        border_top = workbook.add_format({
            'align': 'right',
            'top': 1
        })
        border_top_dec = workbook.add_format({
            'num_format': '#,##0.00_);(#,##0.00)',
            'align': 'right',
            'top': 1
        })
        border_top_num = workbook.add_format({
            'num_format': '#,##0_);(#,##0)',
            'align': 'right',
            'top': 1
        })
        border_top_bot_dec = workbook.add_format({
            'num_format': '#,##0.00_);(#,##0.00)',
            'align': 'right',
            'bottom': 1,
            'top': 1
        })
        border_top_bot_num = workbook.add_format({
            'num_format': '#,##0_);(#,##0)',
            'align': 'right',
            'bottom': 1,
            'top': 1
        })

        dec_format = workbook.add_format({
            'num_format': '#,##0.00_);(#,##0.00)',
            'align': 'right', 
        })
        num_format = workbook.add_format({
            'num_format': '#,##0_);(#,##0)',
            'align': 'right', 
        })

        array_data = str(issue_date).split('-')
        month = int(array_data[1])
        year = int(array_data[0])
        issue_from = datetime.date(int(array_data[0]), int(array_data[1]), 1)
        issue_to = datetime.date(int(array_data[0]), int(array_data[1]),
                                 calendar.monthrange(int(array_data[0]), int(array_data[1]))[1])
        company = Company.objects.get(pk=company_id)
        fsc_calendar = FiscalCalendar.objects.filter(company_id=company_id, is_hidden=0, period=month, fiscal_year=year).last()
        if fsc_calendar:
            end_date = fsc_calendar.end_date.strftime('%d/%m/%Y')
        else:
            end_date = issue_to.strftime('%d/%m/%Y')

        segment_ids = []
        segment_name = ''
        if filter_type == 'undefined' or filter_type == '' or filter_type == 'None':
            filter_type = 'Account'
        if filter_type == 'Segment':
            try:
                from_val = int(from_val)
                to_val = int(to_val)
                segments = CostCenters.objects.filter(id__range=(from_val, to_val))
                for seg in segments:
                    segment_name += ' - ' + seg.name
                segment_ids = segments.values_list('id', flat=True)
            except:
                pass

        worksheet.set_column(0, 0, 12)
        worksheet.set_column(1, 1, 30)
        worksheet.set_column(2, 2, 20)

        decimal_place = "%.2f"
        is_decimal = True
        if not company.currency.is_decimal:
            decimal_place = "%.0f"
            is_decimal = False

        accountHistory = AccountHistory.objects.select_related('account').filter(company_id=company.id, is_hidden=0,
                                                                                 period_month__exact=month,
                                                                                 period_year__exact=year)\
            .exclude(source_currency_id__isnull=True)
        if company.bs_template:
            try:
                bytefile = company.bs_template.read()
                wb = load_workbook(ContentFile(bytefile))
                sheet_0 = wb.get_sheet_names()[0]
                ws = wb[sheet_0]

                hidden_rows = []
                for rowNum, rowDimension in ws.row_dimensions.items():
                    if rowDimension.hidden == True:
                        hidden_rows.append(rowNum)
                hidden_cols = []
                for colNum, colDimension in ws.column_dimensions.items():
                    if colDimension.hidden == True:
                        hidden_cols.append(colNum)

                start = False
                negative_start = False
                total_account = 0
                seg_codes = list(CostCenters.objects.filter(is_active=1, is_hidden=0, company_id=company_id).values_list('code', flat=True))
                account_item_list = Account.objects.filter(company_id=company.id, is_hidden=0).exclude(deactivate_period__lte=issue_from)
                account_item_list = account_item_list.exclude(Q(is_active=False) & Q(deactivate_date=None))
                value_array = [0] * (ws.max_row + 5)
                title_pattern = '[a-zA-Z]+[,]*[a-zA-Z]+[:]*[^0-9]+'
                try:
                    for row in range(1, ws.max_row + 5):
                        empty_row = True
                        if row not in hidden_rows:
                            for col in range(1, ws.max_column + 5):
                                if col not in hidden_cols:
                                    if ws.cell(row, col).value is not None:
                                        if re.search(str(company.currency.code).lower(), str(ws.cell(row, col).value).lower()):
                                            start = True
                                            empty_row = False
                                            printing_row += 1
                                            worksheet.merge_range('A' + str(printing_row) + ':' + 'C' + str(printing_row),
                                                                  str(company.name) + str(segment_name), merge_format)
                                            printing_row += 1
                                            worksheet.merge_range('A' + str(printing_row) + ':' + 'C' + str(printing_row), 'BALANCE SHEET', merge_format)
                                            printing_row += 1
                                            worksheet.merge_range('A' + str(printing_row) + ':' + 'C' + str(printing_row), 'AS AT ' + end_date, merge_format)
                                            printing_row += 1
                                            worksheet.write(printing_row, 2, company.currency.code, center_line)
                                            printing_row += 1
                                            break
                                        if re.search('BALP'.lower(), str(ws.cell(row, col).value).lower()) or \
                                            re.search('CONAME'.lower(), str(ws.cell(row, col).value).lower()) or \
                                                re.search('BALANCE SHEET'.lower(), str(ws.cell(row, col).value).lower()) or \
                                        re.search('AS AT'.lower(), str(ws.cell(row, col).value).lower()):
                                            break
                                        if start:
                                            if re.search(title_pattern, str(ws.cell(row, col).value).replace(' ', '').lower()) or \
                                                    re.search(sum_pattern_1, str(ws.cell(row, col).value)) or re.search(sum_pattern_2, str(ws.cell(row, col).value)):
                                                empty_row = False
                                                if re.search(sum_pattern_1, str(ws.cell(row, col).value)) or re.search(sum_pattern_2, str(ws.cell(row, col).value)):
                                                    result = get_result(str(ws.cell(row, col).value), value_array)
                                                    self.set_cell_style((ws.cell(row, col)))
                                                    if is_decimal:
                                                        local_border = workbook.add_format({
                                                            'num_format': '#,##0.00_);(#,##0.00)',
                                                            'align': 'right'
                                                        })
                                                    else:
                                                        local_border = workbook.add_format({
                                                            'num_format': '#,##0_);(#,##0)',
                                                            'align': 'right'
                                                        })
                                                    
                                                    if self.top_style == 'thin':
                                                        local_border.set_top(1)
                                                    elif self.top_style == 'double':
                                                        local_border.set_top(6)
                                                    if self.bottom_style == 'thin':
                                                        local_border.set_bottom(1)
                                                    elif self.bottom_style == 'double':
                                                        local_border.set_bottom(6)

                                                    value_array[row] = float(result)
                                                    worksheet.write(printing_row, printing_col + 2, wrap_text_xls(result, decimal_place), local_border)
                                                    printing_row += 1
                                                elif re.search(title_pattern, str(ws.cell(row, col).value).replace(' ', '').lower()):
                                                    if re.search('FIXEDASSETS,ATCOST'.lower(), str(ws.cell(row, col).value).replace(' ', '').lower()):
                                                        negative_start = False
                                                    if re.search('INVESTMENT'.lower(), str(ws.cell(row, col).value).replace(' ', '').lower()) or \
                                                            re.search('TOTALFIXEDASSETS'.lower(), str(ws.cell(row, col).value).replace(' ', '').lower()) or \
                                                            re.search('CURRENTASSETS'.lower(), str(ws.cell(row, col).value).replace(' ', '').lower()):
                                                        negative_start = True
                                                    total_str = str(ws.cell(row, col).value).replace('=', '').replace('"', '')
                                                    col += 1
                                                    while ws.cell(row, col).value is None and col < ws.max_column:
                                                        col += 1
                                                    if re.search(sum_pattern_1, str(ws.cell(row, col).value)) or re.search(sum_pattern_2, str(ws.cell(row, col).value)):
                                                        result = get_result(str(ws.cell(row, col).value), value_array)
                                                        self.set_cell_style((ws.cell(row, col)))
                                                        if is_decimal:
                                                            local_border = workbook.add_format({
                                                                'num_format': '#,##0.00_);(#,##0.00)',
                                                                'align': 'right'
                                                            })
                                                        else:
                                                            local_border = workbook.add_format({
                                                                'num_format': '#,##0_);(#,##0)',
                                                                'align': 'right'
                                                            })
                                                        if self.top_style == 'thin':
                                                            local_border.set_top(1)
                                                        elif self.top_style == 'double':
                                                            local_border.set_top(6)
                                                        if self.bottom_style == 'thin':
                                                            local_border.set_bottom(1)
                                                        elif self.bottom_style == 'double':
                                                            local_border.set_bottom(6)

                                                        worksheet.merge_range('A' + str(printing_row + 1) + ':' + 'B' + str(printing_row + 1), total_str, merge_str)
                                                        worksheet.write(printing_row, printing_col + 2, wrap_text_xls(result, decimal_place), local_border)
                                                        printing_row += 1
                                                        value_array[row] = float(result)
                                                    else:
                                                        worksheet.merge_range('A' + str(printing_row + 1) + ':' + 'B' + str(printing_row + 1), total_str, merge_str)
                                                        printing_row += 1
                                                break
                                            elif re.search(acc_pattern_4, str(ws.cell(row, col).value)):  # "2000:2100,2102%%"
                                                empty_row = False
                                                acc_string = str(ws.cell(row, col).value)
                                                col += 1
                                                while ws.cell(row, col).value is None or str(ws.cell(row, col).value) == 'D(ACCT)':
                                                    col += 1
                                                if re.search(sum_pattern_3, str(ws.cell(row, col).value)):
                                                    val_1 = value_array[row - 1]
                                                    val_2 = value_array[row - 2]
                                                    result = val_1 + val_2
                                                    value_array[row] = float(result)
                                                    worksheet.write(printing_row, printing_col + 2, wrap_text_xls(
                                                        result, decimal_place), border_top_bot_dec if is_decimal else border_top_bot_num)
                                                    printing_row += 1
                                                else:
                                                    acc_str_array = acc_string.split(',')
                                                    acc_ids = []
                                                    for acc_str in acc_str_array:
                                                        if re.search('[A-Z]*[0-9]+[:][A-Z]*[0-9]+[Z]*', acc_str):
                                                            acc_range = acc_str
                                                            acc_range = acc_range.replace('Z', '9').replace('=', '').replace('"', '').strip().split(':')
                                                            f_code = int(re.sub('[A-Z]+', '', acc_range[0]))
                                                            l_code = int(re.sub('[A-Z]+', '', acc_range[1]))
                                                            code_list = []
                                                            while f_code <= l_code:
                                                                code_list.append(str(f_code))
                                                                f_code += 1
                                                            ids = account_item_list.filter(account_segment__in=code_list).order_by('id').values_list('id', flat=True)
                                                            acc_ids = acc_ids + list(ids)
                                                        elif re.search('[A-Z]*[0-9]+[%]*', acc_str):
                                                            f_code = acc_str.replace('%', '').replace('=', '').replace('"', '').strip()
                                                            f_code = re.sub('[A-Z]+', '', f_code)
                                                            ids = account_item_list.filter(account_segment=f_code).order_by('id').values_list('id', flat=True)
                                                            acc_ids = acc_ids + list(ids)

                                                    account_list = account_item_list.filter(id__in=acc_ids).order_by('account_segment', 'code')
                                                    if filter_type == 'Segment':
                                                        account_list = account_list.filter(segment_code_id__in=segment_ids)
                                                    if re.search('ACCTDESC'.lower(), str(ws.cell(row, col).value).lower()):
                                                        accumulate = False
                                                    else:
                                                        accumulate = True
                                                    if re.search('FORTHEPERIOD'.lower(), str(ws.cell(row, col).value).replace(' ', '').lower()):
                                                        profit_loss = True
                                                    else:
                                                        profit_loss = False
                                                    name_str = str(ws.cell(row, col).value)

                                                    col += 1
                                                    while not re.search(sum_pattern_3, str(ws.cell(row, col).value)) and col < ws.max_column:
                                                        col += 1
                                                    sign = Decimal(1.0)
                                                    if re.search('\-', str(ws.cell(row, col).value)) and negative_start:
                                                        sign = Decimal(-1.0)
                                                    if report_type == dict(REPORT_TYPE)['Provisional'] and profit_loss:
                                                        sign = Decimal(1.0)

                                                    code_list = account_list.order_by('account_segment').values_list('account_segment', flat=True).distinct()
                                                    accumulate_total = 0
                                                    for cod in code_list:
                                                        total_account = 0
                                                        segments = account_list.filter(account_segment=cod)
                                                        for f_acc in segments:
                                                            item_account_q = accountHistory.filter(account_id=f_acc.id)
                                                            item_account = item_account_q.filter(
                                                                Q(source_currency_id=company.currency_id) |
                                                                Q(source_currency_id__isnull=True))
                                                            if item_account:
                                                                sum_debit = sum_credit = 0
                                                                for iAccount in item_account:
                                                                    if f_acc.balance_type == BALANCE_TYPE_DICT['Credit']:
                                                                        sum_credit += (iAccount.functional_end_balance * sign)
                                                                    if f_acc.balance_type == BALANCE_TYPE_DICT['Debit']:
                                                                        sum_debit += (iAccount.functional_end_balance * sign)
                                                                total_segment = sum_debit + sum_credit
                                                                if report_type == dict(REPORT_TYPE)['Provisional']:
                                                                    prov_posted = getProvisionalPostedAmount(company_id, int(array_data[0]), int(array_data[1]),
                                                                                                             f_acc.id, iAccount.source_currency_id)
                                                                    total_segment += Decimal(prov_posted)
                                                                total_account += total_segment
                                                                accumulate_total += total_segment

                                                                if not accumulate:  # One segment
                                                                    if filter_type == 'Segment':
                                                                        worksheet.write(printing_row, printing_col, f_acc.code)
                                                                        worksheet.write(printing_row, printing_col + 1, f_acc.name)
                                                                        worksheet.write(printing_row, printing_col + 2,
                                                                                        wrap_text_xls(total_segment, decimal_place), dec_format if is_decimal else num_format)
                                                                        printing_row += 1

                                                        if not accumulate:  # All segment
                                                            if filter_type == 'Account':
                                                                f_acc = account_item_list.filter(account_segment=str(cod)).first()
                                                                if f_acc:
                                                                    code = f_acc.code.split('-')[0]
                                                                    if company.use_segment:
                                                                        names = f_acc.name.split('-')
                                                                        for cd in seg_codes:
                                                                            if names[-1] in cd:
                                                                                names.pop()
                                                                        name = '-'.join(names)
                                                                    else:
                                                                        name = f_acc.name
                                                                    worksheet.write(printing_row, printing_col, code)
                                                                    worksheet.write(printing_row, printing_col + 1, name)
                                                                    worksheet.write(printing_row, printing_col + 2,
                                                                                    wrap_text_xls(total_account, decimal_place), dec_format if is_decimal else num_format)
                                                                    printing_row += 1
                                                    if report_type == dict(REPORT_TYPE)['Provisional'] and profit_loss:
                                                        accumulate_total *= -1
                                                    value_array[row] = float(accumulate_total)
                                                    if accumulate:  # accumulate_pattern
                                                        if re.search(accumulate_pattern, name_str):
                                                            name = re.sub(accumulate_pattern, '', name_str).replace('=', '').replace('"', '').strip()
                                                            code = re.sub(name, '', name_str).replace('=', '').replace('"', '').replace(' ', '').strip()
                                                        else:
                                                            code = name_str.replace('=', '').replace('"', '').strip()
                                                            name = ''

                                                        worksheet.write(printing_row, printing_col, code)
                                                        worksheet.write(printing_row, printing_col + 1, name)
                                                        worksheet.write(printing_row, printing_col + 2,
                                                                        wrap_text_xls(accumulate_total, decimal_place), dec_format if is_decimal else num_format)
                                                        printing_row += 1
                                                break
                                            elif re.search(acc_pattern_1, str(ws.cell(row, col).value)):
                                                empty_row = False
                                                acc_range = str(ws.cell(row, col).value)
                                                acc_range = acc_range.replace('Z', '9').replace('=', '').replace('"', '').strip().split(':')
                                                f_code = int(re.sub('[A-Z]+', '', acc_range[0]))
                                                l_code = int(re.sub('[A-Z]+', '', acc_range[1]))
                                                code_list = []
                                                while f_code <= l_code:
                                                    code_list.append(str(f_code))
                                                    f_code += 1
                                                account_list = account_item_list.filter(account_segment__in=code_list).order_by('account_segment', 'code')
                                                if filter_type == 'Segment':
                                                    account_list = account_list.filter(segment_code_id__in=segment_ids)
                                                col += 1
                                                while ws.cell(row, col).value is None or str(ws.cell(row, col).value) == 'D(ACCT)':
                                                    col += 1
                                                if re.search('ACCTDESC'.lower(), str(ws.cell(row, col).value).lower()):
                                                    accumulate = False
                                                else:
                                                    accumulate = True
                                                if re.search('FORTHEPERIOD'.lower(), str(ws.cell(row, col).value).replace(' ', '').lower()):
                                                    profit_loss = True
                                                else:
                                                    profit_loss = False
                                                name_str = str(ws.cell(row, col).value)

                                                col += 1
                                                while not re.search(sum_pattern_3, str(ws.cell(row, col).value)) and col < ws.max_column:
                                                    col += 1
                                                sign = Decimal(1.0)
                                                if re.search('\-', str(ws.cell(row, col).value)) and negative_start:
                                                    sign = Decimal(-1.0)
                                                if report_type == dict(REPORT_TYPE)['Provisional'] and profit_loss:
                                                    sign = Decimal(1.0)

                                                code_list = account_list.order_by('account_segment').values_list('account_segment', flat=True).distinct()
                                                accumulate_total = 0
                                                for cod in code_list:
                                                    total_account = 0
                                                    segments = account_list.filter(account_segment=cod)
                                                    for f_acc in segments:
                                                        item_account_q = accountHistory.filter(account_id=f_acc.id)
                                                        item_account = item_account_q.filter(
                                                            Q(source_currency_id=company.currency_id) |
                                                            Q(source_currency_id__isnull=True))
                                                        if item_account:
                                                            sum_debit = sum_credit = 0
                                                            for iAccount in item_account:
                                                                if f_acc.balance_type == BALANCE_TYPE_DICT['Credit']:
                                                                    sum_credit += (iAccount.functional_end_balance * sign)
                                                                if f_acc.balance_type == BALANCE_TYPE_DICT['Debit']:
                                                                    sum_debit += (iAccount.functional_end_balance * sign)
                                                            total_segment = sum_debit + sum_credit
                                                            if report_type == dict(REPORT_TYPE)['Provisional']:
                                                                prov_posted = getProvisionalPostedAmount(company_id, int(array_data[0]), int(array_data[1]),
                                                                                                         f_acc.id, iAccount.source_currency_id)
                                                                total_segment += Decimal(prov_posted)
                                                            total_account += total_segment
                                                            accumulate_total += total_segment

                                                            if not accumulate:  # One segment
                                                                if filter_type == 'Segment':
                                                                    worksheet.write(printing_row, printing_col, f_acc.code)
                                                                    worksheet.write(printing_row, printing_col + 1, f_acc.name)
                                                                    worksheet.write(printing_row, printing_col + 2,
                                                                                    wrap_text_xls(total_segment, decimal_place), dec_format if is_decimal else num_format)
                                                                    printing_row += 1

                                                    if not accumulate:  # All segment
                                                        if filter_type == 'Account':
                                                            f_acc = account_item_list.filter(account_segment=str(cod)).first()
                                                            if f_acc:
                                                                code = f_acc.code.split('-')[0]
                                                                if company.use_segment:
                                                                    names = f_acc.name.split('-')
                                                                    for cd in seg_codes:
                                                                        if names[-1] in cd:
                                                                            names.pop()
                                                                    name = '-'.join(names)
                                                                else:
                                                                    name = f_acc.name
                                                                worksheet.write(printing_row, printing_col, code)
                                                                worksheet.write(printing_row, printing_col + 1, name)
                                                                worksheet.write(printing_row, printing_col + 2,
                                                                                wrap_text_xls(total_account, decimal_place), dec_format if is_decimal else num_format)
                                                                printing_row += 1
                                                if report_type == dict(REPORT_TYPE)['Provisional'] and profit_loss:
                                                    accumulate_total *= -1
                                                value_array[row] = float(accumulate_total)
                                                if accumulate:  # accumulate_pattern
                                                    if re.search(accumulate_pattern, name_str):
                                                        name = re.sub(accumulate_pattern, '', name_str).replace('=', '').replace('"', '').strip()
                                                        code = re.sub(name, '', name_str).replace('=', '').replace('"', '').replace(' ', '').strip()
                                                    else:
                                                        code = name_str.replace('=', '').replace('"', '').strip()
                                                        name = ''

                                                    worksheet.write(printing_row, printing_col, code)
                                                    worksheet.write(printing_row, printing_col + 1, name)
                                                    worksheet.write(printing_row, printing_col + 2, wrap_text_xls(
                                                        accumulate_total, decimal_place), dec_format if is_decimal else num_format)
                                                    printing_row += 1

                                                break
                                            elif re.search(acc_pattern_3, str(ws.cell(row, col).value)):  # "2000%%,2100%%,2102%%"
                                                empty_row = False
                                                account_str = str(ws.cell(row, col).value)
                                                col += 1
                                                while ws.cell(row, col).value is None or str(ws.cell(row, col).value) == 'D(ACCT)':
                                                    col += 1
                                                if re.search(sum_pattern_3, str(ws.cell(row, col).value)):
                                                    val_1 = value_array[row - 1]
                                                    val_2 = value_array[row - 2]
                                                    result = val_1 + val_2
                                                    value_array[row] = float(result)
                                                    worksheet.write(printing_row, printing_col + 2, wrap_text_xls(
                                                        result, decimal_place), border_top_dec if is_decimal else border_top_num)
                                                    printing_row += 1
                                                else:
                                                    name_str = str(ws.cell(row, col).value)
                                                    col += 1
                                                    while not re.search(sum_pattern_3, str(ws.cell(row, col).value)) and col < ws.max_column:
                                                        col += 1
                                                    sign = Decimal(1.0)
                                                    if re.search('\-', str(ws.cell(row, col).value)):
                                                        sign = Decimal(-1.0)
                                                    acc_list = account_str.replace('%', '').replace('=', '').replace('"', '').strip()
                                                    acc_list = re.sub('[A-Z]+', '', acc_list).split(',')
                                                    account_list = account_item_list.filter(account_segment__in=acc_list).order_by('account_segment', 'code')
                                                    if filter_type == 'Segment':
                                                        account_list = account_list.filter(segment_code_id__in=segment_ids)
                                                    accumulate_total = 0
                                                    for f_acc in account_list:
                                                        item_account_q = accountHistory.filter(account_id=f_acc.id)
                                                        item_account = item_account_q.filter(
                                                            Q(source_currency_id=company.currency_id) | Q(source_currency_id__isnull=True))

                                                        if item_account:
                                                            sum_debit = sum_credit = 0
                                                            for iAccount in item_account:
                                                                if f_acc.balance_type == BALANCE_TYPE_DICT['Credit']:
                                                                    sum_credit += (iAccount.functional_end_balance * sign)
                                                                if f_acc.balance_type == BALANCE_TYPE_DICT['Debit']:
                                                                    sum_debit += (iAccount.functional_end_balance * sign)
                                                            total_account = sum_debit + sum_credit
                                                            if report_type == dict(REPORT_TYPE)['Provisional']:
                                                                prov_posted = getProvisionalPostedAmount(company_id, int(array_data[0]), int(array_data[1]),
                                                                                                         f_acc.id, iAccount.source_currency_id)
                                                                total_account += Decimal(prov_posted)
                                                            accumulate_total += total_account

                                                    value_array[row] = float(accumulate_total)
                                                    if re.search(accumulate_pattern, name_str):
                                                        name = re.sub(accumulate_pattern, '', name_str).replace('=', '').replace('"', '').strip()
                                                        code = re.sub(name, '', name_str).replace('=', '').replace('"', '').replace(' ', '').strip()
                                                    else:
                                                        code = name_str.replace('=', '').replace('"', '').strip()
                                                        name = ''
                                                    worksheet.write(printing_row, printing_col, code)
                                                    worksheet.write(printing_row, printing_col + 1, name)
                                                    worksheet.write(printing_row, printing_col + 2, wrap_text_xls(
                                                        accumulate_total, decimal_place), dec_format if is_decimal else num_format)
                                                    printing_row += 1
                                                break
                                            elif re.search(acc_pattern_2, str(ws.cell(row, col).value)):
                                                empty_row = False
                                                f_code = str(ws.cell(row, col).value).replace('%', '').replace('=', '').replace('"', '').strip()
                                                f_code = re.sub('[A-Z]+', '', f_code)
                                                col += 1
                                                while ws.cell(row, col).value is None or str(ws.cell(row, col).value) == 'D(ACCT)':
                                                    col += 1
                                                if re.search(sum_pattern_3, str(ws.cell(row, col).value)):
                                                    result = value_array[row - 1]
                                                    value_array[row] = float(result)
                                                    worksheet.write(printing_row, printing_col + 2, wrap_text_xls(
                                                        result, decimal_place), border_top_bot_dec if is_decimal else border_top_bot_num)
                                                    printing_row += 1
                                                else:
                                                    name_str = str(ws.cell(row, col).value)

                                                    col += 1
                                                    while not re.search(sum_pattern_3, str(ws.cell(row, col).value)) and col < ws.max_column:
                                                        col += 1
                                                    sign = Decimal(1.0)
                                                    if re.search('\-', str(ws.cell(row, col).value)) and negative_start:
                                                        sign = Decimal(-1.0)
                                                    account_list = account_item_list.filter(account_segment=str(f_code))
                                                    if filter_type == 'Segment':
                                                        account_list = account_list.filter(segment_code_id__in=segment_ids)
                                                    segments_total = 0
                                                    if account_list:
                                                        for f_acc in account_list:
                                                            item_account_q = accountHistory.filter(account_id=f_acc.id)
                                                            item_account = item_account_q.filter(
                                                                Q(source_currency_id=company.currency_id) | Q(source_currency_id__isnull=True))

                                                            if item_account:
                                                                sum_debit = sum_credit = 0
                                                                for iAccount in item_account:
                                                                    if f_acc.balance_type == BALANCE_TYPE_DICT['Credit']:
                                                                        sum_credit += (iAccount.functional_end_balance * sign)
                                                                    if f_acc.balance_type == BALANCE_TYPE_DICT['Debit']:
                                                                        sum_debit += (iAccount.functional_end_balance * sign)
                                                                total_account = sum_debit + sum_credit
                                                                if report_type == dict(REPORT_TYPE)['Provisional']:
                                                                    prov_posted = getProvisionalPostedAmount(company_id, int(array_data[0]), int(array_data[1]),
                                                                                                             f_acc.id, iAccount.source_currency_id)
                                                                    total_account += Decimal(prov_posted)
                                                                segments_total += total_account

                                                                if filter_type == 'Segment':
                                                                    if re.search(code_pattern_1, name_str) and re.search(name_pattern, name_str) or \
                                                                            re.search(code_pattern_2, name_str) and re.search(name_pattern, name_str):
                                                                        code = str(f_acc.code)
                                                                        name = str(f_acc.name)
                                                                    elif re.search(code_pattern_1, name_str) or re.search(code_pattern_2, name_str):
                                                                        code = str(f_acc.code)
                                                                        name = re.sub(code_pattern_1, '', name_str)
                                                                        name = re.sub(code_pattern_2, '', name)
                                                                        name = name.replace('&', '').replace('=', '').replace('"', '').strip()
                                                                    elif re.search(str(f_acc.account_segment), name_str):
                                                                        code = str(f_acc.code)
                                                                        name = re.sub(str(f_acc.account_segment), '', name_str)
                                                                        name = name.replace('&', '').replace('=', '').replace('"', '').strip()
                                                                        if re.search('ACCTDESC'.lower(), str(name).lower()) or re.search('NAME'.lower(), str(name).lower()):
                                                                            name = f_acc.name
                                                                    else:
                                                                        code = ''
                                                                        name = name_str.replace('&', '').replace('=', '').replace('"', '')

                                                                    worksheet.write(printing_row, printing_col, code)
                                                                    worksheet.write(printing_row, printing_col + 1, name)
                                                                    worksheet.write(printing_row, printing_col + 2,
                                                                                    wrap_text_xls(total_account, decimal_place), dec_format if is_decimal else num_format)
                                                                    printing_row += 1
                                                        value_array[row] = float(segments_total)
                                                        if filter_type == 'Account':
                                                            s_account = account_item_list.filter(account_segment=str(f_code)).first()
                                                            if re.search(str(f_acc.account_segment), name_str):
                                                                code = f_acc.code.split('-')[0]
                                                                name = re.sub(str(f_acc.account_segment), '', name_str)
                                                                name = name.replace('&', '').replace('=', '').replace('"', '').strip()
                                                                if re.search('ACCTDESC'.lower(), str(name).lower()) or re.search('NAME'.lower(), str(name).lower()):
                                                                    if company.use_segment:
                                                                        names = f_acc.name.split('-')
                                                                        for cd in seg_codes:
                                                                            if names[-1] in cd:
                                                                                names.pop()
                                                                        name = '-'.join(names)
                                                                    else:
                                                                        name = f_acc.name
                                                            else:
                                                                code = s_account.code.split('-')[0]
                                                                if re.search('ACCTDESC'.lower(), name_str.lower()) or re.search('ACCTDESC'.lower(), name_str.lower()):
                                                                    if company.use_segment:
                                                                        names = s_account.name.split('-')
                                                                        for cd in seg_codes:
                                                                            if names[-1] in cd:
                                                                                names.pop()
                                                                        name = '-'.join(names)
                                                                    else:
                                                                        name = s_account.name
                                                                else:
                                                                    name_str = re.sub(code_pattern_1, '', name_str)
                                                                    name = name_str.replace('&', '').replace('=', '').replace('"', '').strip()

                                                            worksheet.write(printing_row, printing_col, code)
                                                            worksheet.write(printing_row, printing_col + 1, name)
                                                            worksheet.write(printing_row, printing_col + 2,
                                                                            wrap_text_xls(segments_total, decimal_place), dec_format if is_decimal else num_format)
                                                            printing_row += 1

                                                break
                                        else:
                                            empty_row = False
                                        # if re.search(str(company.currency.code).lower(), str(ws.cell(row, col).value).lower()):
                                        #     start = True
                                        #     empty_row = False
                                        #     printing_row += 1
                                        #     break

                            if empty_row:
                                printing_row += 1

                except Exception as e:
                    print(e)
                    worksheet.write(printing_row, printing_col, 'ERROR! ON FILE PROCESSING')
                    printing_row += 1
                    worksheet.write(printing_row, printing_col, 'PLEASE MAKE SURE IT IS CORRECT TEMPLATE.')
                    printing_row += 1

            except Exception as e:
                print(e)
                worksheet.write(printing_row, printing_col, 'ERROR! CAN NOT READ THE TEMPLATE.')
                printing_row += 1

        else:  # No template
            worksheet.write(printing_row, printing_col, 'ERROR! THERE IS NO TEMPLATE TO PROCESS.')
            printing_row += 1

        workbook.close()
        xlsx_data = output.getvalue()
        return xlsx_data


def get_result(cell, value_array):
    result = 0.0
    try:
        formula = re.sub(filter_pattern, '', cell.replace('=', '').replace('"', ''))
        # search for pattern like G10:G15
        multi_formula = re.findall(sum_pattern_4, formula)
        if multi_formula:
            # recreate formula
            m_number = [s for s in re.findall(r'\b\d+\b', multi_formula[0])]
            if int(m_number[0]) < int(m_number[1]):
                n = int(m_number[0]) + 1
                while n < int(m_number[1]):
                    m_number.append(str(n))
                    n += 1
                s_number = [str(s) for s in m_number]
                s_number = '+'.join(s_number)
            else:
                s_number = str(m_number[0])
            formula = re.sub(multi_formula[0], s_number, formula)

        # add () to to every number to make formula error free
        row_number = [int(s) for s in re.findall(r'\b\d+\b', formula)]
        for index in row_number:
            formula = re.sub(str(index), '(' + str(index) + ')', formula, 1)

        # replace index with value
        row_number = [int(s) for s in re.findall(r'\b\d+\b', formula)]
        for index in row_number:
            formula = re.sub('\(' + str(index) + '\)', str(value_array[index]), formula, 1)

        # calculate the result
        result = eval(formula)
    except Exception as e:
        print(e)
        result = 0.0

    return result


def wrap_text_xls(value, decimal_place="%.2f"):
    str_value = intcomma(decimal_place % round_number(value)).replace("-", "")
    if value < 0:
        str_value = '(' + str_value + ')'

    return str_value


def getProvisionalPostedAmount(company_id, p_perd_year, p_perd_month, acc_id, curr):
    prov_posted_amount = 0
    provisional_transaction = Transaction.objects.filter(company_id=company_id, is_hidden=False,
                                                         journal__status=int(STATUS_TYPE_DICT['Prov. Posted']),
                                                         journal__perd_year__lte=p_perd_year,
                                                         journal__perd_month__lte=p_perd_month,
                                                         journal__is_hidden=False,
                                                         account_id=acc_id,
                                                         currency_id=curr)
    if provisional_transaction:
        for prov_trx in provisional_transaction:
            if prov_trx.account.balance_type != prov_trx.functional_balance_type:
                prov_posted_amount += (float(prov_trx.total_amount) * -1)
            else:
                prov_posted_amount += float(prov_trx.total_amount)
    return prov_posted_amount
